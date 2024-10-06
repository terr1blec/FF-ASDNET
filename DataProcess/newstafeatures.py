import numpy as np
import os
import pandas as pd
from matplotlib.path import Path
import math
import cv2
import openpyxl
from multiprocessing import Process
import xlwt
from numba import jit

'''
将注视点与感兴趣区域关联，提取基于感兴趣区域的特征

按顺序执行以下步骤：
1.将Fixation数据关联感兴趣区域，生成视觉元素文件   element(paramas,视频编号)
2.计算第一次注视时间特征   Time_to_first_view(paramas,视频编号)
3.计算特征：注视时间、注视次数、注视时间均值、注视时间方差、注视时长-眼睛、注视时长-鼻子、注视时长-嘴巴、注视时长-脸部其他区域、注视时长-视频其他区域、视频中心距均值、视频中心距方差、注视时间均值-眼睛、注视时间均值-鼻子、注视时间均值-嘴巴、
注视时间均值-脸部其他区域、注视时间均值-视频其他区域、注视次数-眼睛、注视次数-鼻子、注视次数-嘴巴、注视次数-脸部其他区域、注视次数-视频其他区域   stastic_features(paramas,视频编号)
4.把两次计算的特征组合在一起   joint(paramas,视频编号)
'''

size = {'media1': 234, 'media2': 286, 'media3':44,'media4':225, 'media5': 186, 'media6':64,'media7': 80,
            'media8': 83, 'media9': 152,'media10': 128, 'media11': 135,'media12':58, 'media13':75,'media14':86}   #每个视频的帧数

class paramas():
    def __init__(self):
        self.datafile = '20210623'
        self.savefile = '1datadecolumn'

        self.savesaccadecoor = '2saccadecoor'
        self.savegazepointcoor = '2gazepointcoor'

        self.savesaccadefixation = '3saccadefixation'
        self.savefixation = '3fixation'
        self.savegazepoint = '3gazepoint'
        self.videostart = 'videostart'

        self.oksaccadefixation = '5oksaccadefixation'
        self.okfixation = '5okfixation'
        self.okgazepoint = '5okgazepoint'
        self.videostartsplit = 'videostartsplit'
        self.videoendsplit = 'videoendsplit'
        self.videostartrowsplit = 'videostartrowsplit'
        self.videoendrowsplit = 'videoendrowsplit'

        self.fixationsample = '6fixationsample'
        self.saccadefixationsample = '6saccadefixationsample'
        self.gazepointsample = '6gazepointsample'

        self.fixationimgcoor = 'fixationimg/fixationcoor'
        self.saccadefixationimgcoor = 'saccadefixationimg/fixationcoor'
        self.gazepointimgcoor = 'gazepointimg/fixationcoor'
        self.gazepointfps = 'gazepointfps'

        self.fixationfsjoint = 'fixationimg/fsjoint14ss'
        self.saccadefixationfsjoint = 'saccadefixationimg/fsjoint10ss'
        self.gazepointfsjoint = 'gazepointimg/fsjoint14ss'
        self.newfixationfsjoint = 'newfixationimg/fsjoint10ss'
        self.newsaccadefixationfsjoint = 'newsaccadefixationimg/fsjoint10ss'
        self.newgazepointfsjoint = 'newgazepointimg/fsjoint10ss'
        self.saliency='../Saliency'
        self.newsaliency='../TASED-Net-master/output'
        self.frame='../img/frame'
        self.gazepointfsjoint1d='gazepointimg/fsjoint1d'
        self.newgazepointfsjoint1d = 'newgazepointimg/fsjoint1d'
        self.fixationfsjoint1d='fixationimg/fsjoint1d'
        self.newfixationfsjoint1d='newfixationimg/fsjoint1d'
        self.saccadefixationfsjoint1d='saccadefixationimg/fsjoint1d'
        self.newsaccadefixationfsjoint1d='newsaccadefixationimg/fsjoint1d'

        self.framecoor='framecoordinate'   #AOI边界点坐标路径
        self.visualelement='visualelement'   #视觉元素保存的文件
        self.visualelementgaze = 'visualelementgaze'
        self.firstfixationtime='stafeatures/firstfixationtime'
        self.aoifeatures='stafeatures/aoi'
        self.aoifeatureswithmeanduration = 'stafeatures/aoiwithduration'
        self.stafeatures='stafeatures/stafeatures'
        self.stafeatureswithmeand = 'stafeatures/stafeatureswithmeand'
        self.stapath = 'stapath'
        self.similarity='stafeatures/similarity'
        self.rateaoifeatures='stafeatures/rateaoifeatures'
        self.ratefirsttime = 'stafeatures/ratefirsttime'   #第一次注视特征保存的文件
        self.ratestafeatures='stafeatures/ratestafeatures'

'''
1.将Fixation数据关联感兴趣区域，生成视觉元素文件
   element(paramas,Fixation样本名)
处理后的文件保存在paramas.visualelement='visualelement'
'''
def element(paramas,file):
    if not os.path.exists(f'{paramas.visualelement}/media{file}'):
        os.makedirs(f'{paramas.visualelement}/media{file}')
    landmark=pd.read_excel(f'{paramas.framecoor}/媒体{file}.xlsx',header=None)
    df=landmark.values.tolist()
    samples=os.listdir(f'{paramas.fixationimgcoor}/media{file}')
    for sample in samples:
        e=0
        individual_path=[]
        fixations=pd.read_excel(f'{paramas.fixationsample}/{sample[:-4]}.xlsx')
        fixations=fixations.values.tolist()
        start=pd.read_excel(f'{paramas.videostartsplit}/{sample[:-4]}.xlsx',header=None)
        start=start.values.tolist()
        for i in range(len(start)):
            if start[i][2][2:-4]==str(file):
                T0=int(start[i][0])
                break

        for co in fixations:
            if co[2][2:-4]==str(file) and str(co[7])!='nan':
                fps = (int(co[0]) - T0) // 40000
                if fps==0 or fps>size['media'+str(file)]:
                    pass
                else:
                    e+=1
                    eye1x=int(df[fps][10])
                    eye1y=int(df[fps][11])
                    eye2x=int(df[fps][12])
                    eye2y=int(df[fps][13])
                    eye3x=int(df[fps][14])
                    eye3y=int(df[fps][15])
                    eye4x=int(df[fps][16])
                    eye4y=int(df[fps][17])
                    eye=Path([(eye1x,eye1y),(eye2x,eye2y),(eye3x,eye3y),(eye4x,eye4y)])
                    nose1x=df[fps][18]
                    nose1y=df[fps][19]
                    nose2x=df[fps][20]
                    nose2y=df[fps][21]
                    nose3x=df[fps][22]
                    nose3y=df[fps][23]
                    nose4x=df[fps][24]
                    nose4y=df[fps][25]
                    nose=Path([(nose1x,nose1y),(nose2x,nose2y),(nose3x,nose3y),(nose4x,nose4y)])
                    mouth1x=df[fps][26]
                    mouth1y=df[fps][27]
                    mouth2x=df[fps][28]
                    mouth2y=df[fps][29]
                    mouth3x=df[fps][30]
                    mouth3y=df[fps][31]
                    mouth4x=df[fps][32]
                    mouth4y=df[fps][33]
                    mouth=Path([(mouth1x,mouth1y),(mouth2x,mouth2y),(mouth3x,mouth3y),(mouth4x,mouth4y)])
                    centerX=df[fps][1]
                    centerY = df[fps][2]
                    short=math.ceil(df[fps][7])
                    long=math.ceil(df[fps][8])
                    angle=df[fps][9]

                    co[7]=round(float(co[7])*720)
                    co[8] = round(float(co[8]) * 576)
                    x=co[7]
                    y=co[8]
                    coo=np.array((x,y))
                # coo[0][0]=x
                # coo[0][1]=y
                    coo.reshape(1,2)
                    if math.pow(((x-centerX)*math.cos(angle)+(y-centerY)*math.sin(angle))/long,2)+math.pow(((centerX-x)*math.sin(angle)+(y-centerY)*math.cos(angle))/short,2)<=1:
                        face_in=True
                    else:
                        face_in=False
                    eye_in = eye.contains_point(coo)
                    nose_in=nose.contains_point(coo)
                    mouth_in=mouth.contains_point(coo)

                    if not(face_in):
                        result='E'
                    elif face_in and not(eye_in or nose_in or mouth_in):
                        result='D'
                    elif face_in and eye_in:
                        if not(nose_in):
                            result='A'
                        else:
                            result='A'
                    elif face_in and nose_in:
                        if not(mouth_in):
                            result='B'
                        else:
                            result='B'
                    elif face_in and mouth_in:
                        result='C'
                    # else:
                    #     result='error'

                    co.append(result)
                    individual_path.append(co)
        if e>1:
            f = open(f'{paramas.visualelement}/media{file}/{sample[:-4]}.txt', 'w')
            f.write('RecordTime'+'\t'+'Participant'+'\t'+'Media'+'\t'+'EyeMovement'+'\t'+'EventDuration'+'\t'+'Duration'+'\t'+'EventIndex'+'\t'+'coordinateX'+'\t'+'coordinateY'+'\t'+'AOI'+'\n')
            for i in range(len(individual_path)):
                for j in range(len(individual_path[i])):
                    f.write(str(individual_path[i][j]))  # write函数不能写int类型的参数，所以使用str()转化
                    f.write('\t')  # 相当于Tab一下，换一个单元格
                f.write('\n')  # 写完一行立马换行
                # f.writelines(str(individual_path))
            f.close()
'''
2.计算第一次注视时间特征
   Time_to_first_view(paramas,视觉元素文件名)
处理后的文件保存在paramas.ratefirsttime = 'stafeatures/ratefirsttime'
'''
def Time_to_first_view(paramas,file):
    if not os.path.exists(f'{paramas.ratefirsttime}'):
        os.makedirs(f'{paramas.ratefirsttime}')
    samples=os.listdir(paramas.fixationsample)
    features=[]
    for sample in samples:
        videostart=pd.read_excel(f'{paramas.videostartsplit}/{sample}',header=None)
        videostart=videostart.values.tolist()
        videoend = pd.read_excel(f'{paramas.videoendsplit}/{sample}', header=None)
        videoend = videoend.values.tolist()
        for i in range(len(videostart)):
            if videostart[i][2][2:-4]==str(file):
                start=int(videostart[i][0])
                break
        for i in range(len(videoend)):
            if videoend[i][2][2:-4]==str(file):
                end=int(videoend[i][0])
                break
        mediatime=end-start
        sheet2=pd.read_excel(f'{paramas.fixationsample}/{sample}')
        sheet2=sheet2.values.tolist()
        for sh in sheet2:
            if sh[2][2:-4]==str(file):
                firstTime=round((int(sh[0])-start)/mediatime,4)
                features.append([sample.split('.')[0],firstTime])
                break

    f = xlwt.Workbook()
    sheet3 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    sheet3.write(0, 0, 'sample')
    sheet3.write(0, 1, 'FirstFixationTime')
    for i in range(len(features)):
        for j in range(len(features[i])):
            sheet3.write(i+1,j,features[i][j])
    f.save(f'{paramas.ratefirsttime}/media{file}.xls')

'''
3.计算特征：注视时间、注视次数、注视时间均值、注视时间方差、注视时长-眼睛、注视时长-鼻子、注视时长-嘴巴、注视时长-脸部其他区域、注视时长-视频其他区域、视频中心距均值、视频中心距方差、注视时间均值-眼睛、注视时间均值-鼻子、注视时间均值-嘴巴、
注视时间均值-脸部其他区域、注视时间均值-视频其他区域、注视次数-眼睛、注视次数-鼻子、注视次数-嘴巴、注视次数-脸部其他区域、注视次数-视频其他区域
   stastic_features(paramas,视频编号)
处理后的文件保存在paramas.aoifeatureswithmeanduration = 'stafeatures/aoiwithduration'
'''
def stastic_features(paramas,file):
    if not os.path.exists(f'{paramas.aoifeatureswithmeanduration}'):
        os.makedirs(f'{paramas.aoifeatureswithmeanduration}')
    samples = os.listdir(f'{paramas.visualelement}/media{file}')
    features = []
    minfixations, minduration, minm_duration, minsd_duration, minfixa, minfixb, minfixc, minfixd, minfixe, mindura, mindurb, mindurc, \
    mindurd, mindure, minmdisimg, minsddisimg, minmdisscan, minsddisscan,minrevisits,minmeanda,minmeandb,minmeandc,minmeandd,minmeande =\
        10e9,10e9,10e9,10e9,10e9,10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9
    maxfixations, maxduration, maxm_duration, maxsd_duration, maxfixa, maxfixb, maxfixc, maxfixd, maxfixe, maxdura, maxdurb, maxdurc, \
    maxdurd, maxdure, maxmdisimg, maxsddisimg, maxmdisscan, maxsddisscan,maxrevisits,maxmeanda,maxmeandb,maxmeandc,maxmeandd,maxmeande =\
        0,0,0,0,0,0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
    for sample in samples:
        fo = open(f'{paramas.visualelement}/media{file}/{sample}', "r")
        myFile = fo.read()
        myRecords = myFile.split('\n')
        myRecords_templist = []
        for y in range(1, len(myRecords) - 1):
            myRecords_templist.append(myRecords[y].split('\t'))  # 把表格数据存入
        duration,m_duration,sd_duration,fixa,fixb,fixc,fixd,fixe,dura,durb,durc,durd,dure,mdisimg,sddisimg,mdisscan,\
            sddisscan,revisits,meanda,meandb,meandc,meandd,meande= 0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0
        disimg,disscan,fixcoor,record = [],[],[],[]
        k = 0  # 当前为第k条记录
        m = 0  # m=0表示
        fixations=len(myRecords_templist)
        for data in myRecords_templist:
            duration+=int(data[5])
            record.append(int(data[5]))
            fixcoor.append([float(data[7]),float(data[8])])
            if data[9] == 'A':
                fixa += 1
                dura += int(data[5])
            elif data[9] == 'B':
                fixb += 1
                durb += int(data[5])
            elif data[9] == 'C':
                fixc += 1
                durc += int(data[5])
            elif data[9] == 'D':
                fixd += 1
                durd += int(data[5])
            elif data[9] == 'E':
                fixe += 1
                dure += int(data[5])
        for j in range(len(fixcoor)):
            disimg.append(math.sqrt(math.pow((fixcoor[j][0] - 360), 2) + math.pow((fixcoor[j][1] - 288), 2)))
        mdisimg = np.mean(disimg)
        sddisimg = np.var(disimg)
        if fixa>1:
            revisits=revisits+fixa-1
        if fixb>1:
            revisits=revisits+fixb-1
        if fixc>1:
            revisits=revisits+fixc-1
        if fixd>1:
            revisits=revisits+fixd-1
        if fixe>1:
            revisits=revisits+fixe-1
        if fixa>0:
            meanda=dura/fixa
        if fixb>0:
            meandb = durb / fixb
        if fixc>0:
            meandc = durc / fixc
        if fixd>0:
            meandd = durd / fixd
        if fixe>0:
            meande = dure / fixe
        if len(fixcoor) % 2 == 0:
            index = int(len(fixcoor) / 2) - 1
            x = (fixcoor[index][0] + fixcoor[index + 1][0]) / 2
            y = (fixcoor[index][1] + fixcoor[index + 1][1]) / 2
            for j in range(len(fixcoor)):
                disscan.append(math.sqrt(math.pow((fixcoor[j][0] - x), 2) + math.pow((fixcoor[j][1] - y), 2)))
            mdisscan = np.mean(disscan)
            sddisscan = np.var(disscan)
        else:
            if len(fixcoor) == 1:
                mdisscan = 0
                sddisscan = 0
            else:
                index = len(fixcoor) // 2
                x = fixcoor[index][0]
                y = fixcoor[index][1]
                for j in range(len(fixcoor)):
                    disscan.append(
                        math.sqrt(math.pow((fixcoor[j][0] - x), 2) + math.pow((fixcoor[j][1] - y), 2)))
                mdisscan = np.mean(disscan)
                sddisscan = np.var(disscan)
        m_duration = np.mean(record)
        sd_duration = np.var(record)

        if fixations>maxfixations:
            maxfixations=fixations
        if fixations<minfixations:
            minfixations=fixations
        if duration > maxduration:
            maxduration = duration
        if duration < minduration:
            minduration = duration
        if m_duration>maxm_duration:
            maxm_duration=m_duration
        if m_duration<minm_duration:
            minm_duration=m_duration
        if sd_duration>maxsd_duration:
            maxsd_duration=sd_duration
        if sd_duration<minsd_duration:
            minsd_duration=sd_duration
        if fixa>maxfixa:
            maxfixa=fixa
        if fixa<minfixa:
            minfixa=fixa
        if fixb > maxfixb:
            maxfixb = fixb
        if fixb < minfixb:
            minfixb = fixb
        if fixc > maxfixc:
            maxfixc = fixc
        if fixc < minfixc:
            minfixc= fixc
        if fixd > maxfixd:
            maxfixd = fixd
        if fixd < minfixd:
            minfixd = fixd
        if fixe > maxfixe:
            maxfixe = fixe
        if fixe < minfixe:
            minfixe = fixe
        if dura>maxdura:
            maxdura=dura
        if dura<mindura:
            mindura=dura
        if durb > maxdurb:
            maxdurb = durb
        if durb < mindurb:
            mindurb = durb
        if durc > maxdurc:
            maxdurc = durc
        if durc < mindurc:
            mindurc= durc
        if durd > maxdurd:
            maxdurd = durd
        if durd < mindurd:
            mindurd = durd
        if dure > maxdure:
            maxdure = dure
        if dure < mindure:
            mindure = dure
        if mdisimg>maxmdisimg:
            maxmdisimg=mdisimg
        if mdisimg<minmdisimg:
            minmdisimg=mdisimg
        if sddisimg>maxsddisimg:
            maxsddisimg=sddisimg
        if sddisimg<minsddisimg:
            minsddisimg=sddisimg
        if mdisscan>maxmdisscan:
            maxmdisscan=mdisscan
        if mdisscan<minmdisscan:
            minmdisscan=mdisscan
        if sddisscan>maxsddisscan:
            maxsddisscan=sddisscan
        if sddisscan<minsddisscan:
            minsddisscan=sddisscan
        if revisits>maxrevisits:
            maxrevisits=revisits
        if revisits<minrevisits:
            minrevisits=revisits
        if meanda>maxmeanda:
            maxmeanda=meanda
        if meandb>maxmeandb:
            maxmeandb=meandb
        if meandc>maxmeandc:
            maxmeandc=meandc
        if meandd>maxmeandd:
            maxmeandd=meandd
        if meande>maxmeande:
            maxmeande=meande
        if meanda<minmeanda:
            minmeanda=meanda
        if meandb<minmeandb:
            minmeandb=meandb
        if meandc<minmeandc:
            minmeandc=meandc
        if meandd<minmeandd:
            minmeandd=meandd
        if meande<minmeande:
            minmeande=meande

        features.append([sample.split('.')[0],fixations,duration,m_duration,sd_duration,fixa,fixb,fixc,fixd,fixe,dura,durb,durc,durd,
                         dure,mdisimg,sddisimg,mdisscan,sddisscan,revisits,meanda,meandb,meandc,meandd,meande])
    f = xlwt.Workbook()
    sheet3 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    line1=['sample','fixations','duration','m_duration','sd_duration','fixa','fixb','fixc','fixd','fixe','dura','durb','durc','durd',
           'dure','mdisimg','sddisimg','mdisscan','sddisscan','revisits','meanda','meandb','meandc','meandd','meande','gyhfixations','gyhduration','gyhm_duration','gyhsd_duration','gyhfixa',
           'gyhfixb','gyhfixc','gyhfixd','gyhfixe','gyhdura','gyhdurb','gyhdurc','gyhdurd','gyhdure',
           'gyhmdisimg','gyhsddisimg','gyhmdisscan','gyhsddisscan','gyhrevisits','gyhmeanda','gyhmeandb','gyhmeandc','gyhmeandd','gyhmeande']
    for i in range(len(line1)):
        sheet3.write(0,i,line1[i])

    for i in range(len(features)):
        gyhfixations=np.round((features[i][1]-minfixations)/(maxfixations-minfixations),4)
        gyhduration = np.round((features[i][2] - minduration) / (maxduration - minduration), 4)
        gyhm_duration = np.round((features[i][3] - minm_duration) / (maxm_duration - minm_duration), 4)
        gyhsd_duration = np.round((features[i][4] - minsd_duration) / (maxsd_duration - minsd_duration), 4)
        gyhfixa = np.round((features[i][5] - minfixa) / (maxfixa - minfixa), 4)
        gyhfixb = np.round((features[i][6] - minfixb) / (maxfixb - minfixb), 4)
        gyhfixc = np.round((features[i][7] - minfixc) / (maxfixc - minfixc), 4)
        gyhfixd = np.round((features[i][8] - minfixd) / (maxfixd - minfixd), 4)
        gyhfixe = np.round((features[i][9] - minfixe) / (maxfixe - minfixe), 4)
        gyhdura = np.round((features[i][10] - mindura) / (maxdura - mindura), 4)
        gyhdurb = np.round((features[i][11] - mindurb) / (maxdurb - mindurb), 4)
        gyhdurc = np.round((features[i][12] - mindurc) / (maxdurc - mindurc), 4)
        gyhdurd = np.round((features[i][13] - mindurd) / (maxdurd - mindurd), 4)
        gyhdure = np.round((features[i][14] - mindure) / (maxdure - mindure), 4)
        gyhmdisimg = np.round((features[i][15] - minmdisimg) / (maxmdisimg - minmdisimg), 4)
        gyhsddisimg = np.round((features[i][16] - minsddisimg) / (maxsddisimg - minsddisimg), 4)
        gyhmdisscan = np.round((features[i][17] - minmdisscan) / (maxmdisscan - minmdisscan), 4)
        gyhsddisscan = np.round((features[i][18] - minsddisscan) / (maxsddisscan - minsddisscan), 4)
        gyhrevisits = np.round((features[i][19] - minrevisits) / (maxrevisits - minrevisits), 4)
        gyhmeanda = np.round((features[i][20] - minmeanda) / (maxmeanda - minmeanda), 4)
        gyhmeandb = np.round((features[i][21] - minmeandb) / (maxmeandb - minmeandb), 4)
        gyhmeandc = np.round((features[i][22] - minmeandc) / (maxmeandc - minmeandc), 4)
        gyhmeandd = np.round((features[i][23] - minmeandd) / (maxmeandd - minmeandd), 4)
        gyhmeande = np.round((features[i][24] - minmeande) / (maxmeande - minmeande), 4)
        features[i].extend([gyhfixations,gyhduration,gyhm_duration,gyhsd_duration,gyhfixa,gyhfixb,gyhfixc,gyhfixd,
                            gyhfixe,gyhdura,gyhdurb,gyhdurc,gyhdurd,gyhdure,gyhmdisimg,gyhsddisimg,gyhmdisscan,gyhsddisscan,gyhrevisits,
                            gyhmeanda,gyhmeandb,gyhmeandc,gyhmeandd,gyhmeande])
        for j in range(len(features[i])):
            sheet3.write(i+1,j,features[i][j])
    f.save(f'{paramas.aoifeatureswithmeanduration}/media{file}.xls')

def joint(paramas,file):
    features=pd.read_excel(f'{paramas.rateaoifeatures}/media{file}.xls')
    features=features.values.tolist()
    firsttime=pd.read_excel(f'{paramas.ratefirsttime}/media{file}.xls')
    firsttime=firsttime.values.tolist()
    f = xlwt.Workbook()
    sheet3 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    line1 = ['sample', 'duration','m_duration','sd_duration','dura','durb','durc','durd',
           'dure','mdisimg','sddisimg','meanda','meandb','meandc','meandd','meande','gduration','gdura','gdurb','gdurc','gdurd','gdure','firsttime']
    for i in range(len(line1)):
        sheet3.write(0,i,line1[i])
    for i in range(len(features)):
        final=[]
        final.extend(features[i])
        for j in range(len(firsttime)):
            if firsttime[j][0]==features[i][0]:
                final.append(firsttime[j][1])
                break
        for j in range(len(final)):
            sheet3.write(i+1,j,final[j])
    f.save(f'{paramas.ratestafeatures}/media{file}.xls')

def spssdata(pa):
    fo = open(f'../sample/20220817wos4/14/0/trainvalidationsample.txt', 'r')
    train = []
    myFile = fo.read()
    myRecords = myFile.split('\n')
    for y in range(0, len(myRecords) - 1):
        train.append(myRecords[y].split('\t'))  # 把表格数据存入
    fo1 = open(f'../sample/20220817wos4/14/0/testsample.txt', 'r')
    myFile1 = fo1.read()
    myRecords1 = myFile1.split('\n')
    for y in range(0, len(myRecords1) - 1):
        train.append(myRecords1[y].split('\t'))  # 把表格数据存入
    print(len(train))
    features={}
    for i in range(1,15):
        fea = pd.read_excel(f'{pa.ratestafeatures}/media{i}.xls')
        features['media'+str(i)] = fea.values.tolist()

    data = []
    data.append(
        ['Participant', 'isASD', 'media', 'happy', 'duration','m_duration','sd_duration','dura','durb','durc','durd',
           'dure','mdisimg','sddisimg','meanda','meandb','meandc','meandd','meande','gduration','gdura','gdurb','gdurc','gdurd','gdure','firsttime'])
    for i in range(len(train)):
        e=0
        re=[]
        if train[i][1][0] == 'n':
            isasd = 2
        else:
            isasd = 1
        if int(train[i][0][5:]) > 7:
            happy = 1
        else:
            happy = 2
        re.extend([train[i][1],isasd,int(train[i][0][5:]),happy])
        for feature in features[train[i][0]]:
            if feature[0]==train[i][1]:
                re.extend(feature[1:])
                data.append(re)
                e=1
                break
        if e==0:
            print(f'{train[i][0]}  {train[i][1]}')

    f = xlwt.Workbook()
    sheet3 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    for i in range(len(data)):
        for j in range(len(data[i])):
            sheet3.write(i, j, data[i][j])
    f.save(f'spssdata17rate.xls')


pa=paramas()
if not os.path.exists(f'{pa.rateaoifeatures}'):
    os.makedirs(f'{pa.rateaoifeatures}')
if not os.path.exists(f'{pa.ratestafeatures}'):
    os.makedirs(f'{pa.ratestafeatures}')
process_list = []
# for i in range(1,15):  #开启5个子进程执行fun1函数
#     p = Process(target=joint,args=(pa,i)) #实例化进程对象
#     p.start()
#     process_list.append(p)
#
# for i in process_list:
#     i.join()
# spssdata(pa)




'''
def gaze_element(paramas,file):
    if not os.path.exists(f'{paramas.visualelementgaze}/media{file}'):
        os.makedirs(f'{paramas.visualelementgaze}/media{file}')
    landmark=pd.read_excel(f'{paramas.framecoor}/媒体{file}.xlsx',header=None)
    df=landmark.values.tolist()
    samples=os.listdir(f'{paramas.gazepointsample}')

    for sample in samples:
        mz=sample[:-5]
        timestart=pd.read_excel(f'{paramas.videostartrowsplit}/{sample[:-5]}.xlsx',header=None)
        timestart=timestart.values.tolist()
        timeend = pd.read_excel(f'{paramas.videoendrowsplit}/{sample[:-5]}.xlsx',header=None)
        timeend=timeend.values.tolist()
        starts = pd.read_excel(f'{paramas.videostartsplit}/{sample[:-5]}.xlsx', header=None)
        starts = starts.values.tolist()
        for i in range(len(timestart)):
            if timestart[i][2][2:-4]==str(file):
                mediastart=int(timestart[i][0])
                break
        for i in range(len(timeend)):
            if timeend[i][2][2:-4]==str(file):
                mediaend=int(timeend[i][0])
                break
        for i in range(len(starts)):
            if starts[i][2][2:-4]==str(file):
                start=int(starts[i][0])
                break
        mediatime=mediaend-mediastart
        if mediatime<=0:
            print(f'media{file} {sample[:-5]} 视频播放时间出错')
        e=0
        ad,bd,cd,dd,ed=0,0,0,0,0
        individual_path=[]
        fixations=pd.read_excel(f'{paramas.gazepointsample}/{sample[:-5]}.xlsx')
        fixations=fixations.values.tolist()

        for i in range(len(fixations)):
            if fixations[i][4][2:-4]==str(file) :
                fps = (int(fixations[i][0]) - start) // 40000
                if fps==0 or fps>size['media'+str(file)]:
                    pass
                else:
                    e+=1
                    eye1x=int(df[fps][10])
                    eye1y=int(df[fps][11])
                    eye2x=int(df[fps][12])
                    eye2y=int(df[fps][13])
                    eye3x=int(df[fps][14])
                    eye3y=int(df[fps][15])
                    eye4x=int(df[fps][16])
                    eye4y=int(df[fps][17])
                    eye=Path([(eye1x,eye1y),(eye2x,eye2y),(eye3x,eye3y),(eye4x,eye4y)])
                    nose1x=df[fps][18]
                    nose1y=df[fps][19]
                    nose2x=df[fps][20]
                    nose2y=df[fps][21]
                    nose3x=df[fps][22]
                    nose3y=df[fps][23]
                    nose4x=df[fps][24]
                    nose4y=df[fps][25]
                    nose=Path([(nose1x,nose1y),(nose2x,nose2y),(nose3x,nose3y),(nose4x,nose4y)])
                    mouth1x=df[fps][26]
                    mouth1y=df[fps][27]
                    mouth2x=df[fps][28]
                    mouth2y=df[fps][29]
                    mouth3x=df[fps][30]
                    mouth3y=df[fps][31]
                    mouth4x=df[fps][32]
                    mouth4y=df[fps][33]
                    mouth=Path([(mouth1x,mouth1y),(mouth2x,mouth2y),(mouth3x,mouth3y),(mouth4x,mouth4y)])
                    centerX=df[fps][1]
                    centerY = df[fps][2]
                    short=math.ceil(df[fps][7])
                    long=math.ceil(df[fps][8])
                    angle=df[fps][9]

                    x=round(float(fixations[i][2])*720)
                    y = round(float(fixations[i][3]) * 576)
                    coo=np.array((x,y))
                # coo[0][0]=x
                # coo[0][1]=y
                    coo.reshape(1,2)
                    if math.pow(((x-centerX)*math.cos(angle)+(y-centerY)*math.sin(angle))/long,2)+math.pow(((centerX-x)*math.sin(angle)+(y-centerY)*math.cos(angle))/short,2)<=1:
                        face_in=True
                    else:
                        face_in=False
                    eye_in = eye.contains_point(coo)
                    nose_in=nose.contains_point(coo)
                    mouth_in=mouth.contains_point(coo)

                    if not(face_in):
                        result='E'
                        ed+=1
                    elif face_in and not(eye_in or nose_in or mouth_in):
                        result='D'
                        dd+=1
                    elif face_in and eye_in:
                        if not(nose_in):
                            result='A'
                            ad+=1
                        else:
                            result='A'
                            ad+=1
                    elif face_in and nose_in:
                        if not(mouth_in):
                            result='B'
                            bd+=1
                        else:
                            result='B'
                            bd+=1
                    elif face_in and mouth_in:
                        result='C'
                        cd+=1

        if ad>0:
            individual_path.append([0,sample[:-5],'媒体'+str(file)+'.avi','Gaze',str(ad/mediatime),str(ad/mediatime),0,0,0,'A'])
        if bd>0:
            individual_path.append([0,sample[:-5],'媒体'+str(file)+'.avi','Gaze',str(bd/mediatime),str(bd/mediatime),0,0,0,'B'])
        if cd>0:
            individual_path.append([0,sample[:-5],'媒体'+str(file)+'.avi','Gaze',str(cd/mediatime),str(cd/mediatime),0,0,0,'C'])
        if dd>0:
            individual_path.append([0,sample[:-5],'媒体'+str(file)+'.avi','Gaze',str(dd/mediatime),str(dd/mediatime),0,0,0,'D'])
        if ed>0:
            individual_path.append([0,sample[:-5],'媒体'+str(file)+'.avi','Gaze',str(ed/mediatime),str(ed/mediatime),0,0,0,'E'])

        f = open(f'{paramas.visualelementgaze}/media{file}/{sample[:-5]}.txt', 'w')
        f.write('RecordTime'+'\t'+'Participant'+'\t'+'Media'+'\t'+'EyeMovement'+'\t'+'EventDuration'+'\t'+'Duration'+'\t'+'EventIndex'+'\t'+'coordinateX'+'\t'+'coordinateY'+'\t'+'AOI'+'\n')
        for i in range(len(individual_path)):
            for j in range(len(individual_path[i])):
                f.write(str(individual_path[i][j]))  # write函数不能写int类型的参数，所以使用str()转化
                f.write('\t')  # 相当于Tab一下，换一个单元格
            f.write('\n')  # 写完一行立马换行
            # f.writelines(str(individual_path))
        f.close()
        
def stastic_features(paramas,file):
    samples = os.listdir(f'{paramas.visualelement}/media{file}')
    features = []
    mediatime=round(size['media'+str(file)]*40)
    for sample in samples:
        fo = open(f'{paramas.visualelement}/media{file}/{sample}', "r")
        myFile = fo.read()
        myRecords = myFile.split('\n')
        myRecords_templist = []
        for y in range(1, len(myRecords) - 1):
            myRecords_templist.append(myRecords[y].split('\t'))  # 把表格数据存入
        fo1 = open(f'{paramas.visualelementgaze}/media{file}/{sample}', "r")
        myFile1 = fo1.read()
        myRecords1 = myFile1.split('\n')
        myRecords_templist1 = []
        for y in range(1, len(myRecords1) - 1):
            myRecords_templist1.append(myRecords1[y].split('\t'))  # 把表格数据存入
        duration,m_duration,sd_duration,dura,durb,durc,durd,dure,mdisimg,sddisimg,meanda,meandb,meandc,meandd,meande,gduration,gdura,gdurb,gdurc,\
        gdurd,gdure,fixa,fixb,fixc,fixd,fixe= 0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0
        disimg,fixcoor,record = [],[],[]
        k = 0  # 当前为第k条记录
        m = 0  # m=0表示
        fixations=len(myRecords_templist)
        for data in myRecords_templist:
            duration+=int(data[5])
            record.append(int(data[5]))
            fixcoor.append([float(data[7]),float(data[8])])
            if data[9] == 'A':
                fixa += 1
                dura += int(data[5])
            elif data[9] == 'B':
                fixb += 1
                durb += int(data[5])
            elif data[9] == 'C':
                fixc += 1
                durc += int(data[5])
            elif data[9] == 'D':
                fixd += 1
                durd += int(data[5])
            elif data[9] == 'E':
                fixe += 1
                dure += int(data[5])
        for j in range(len(fixcoor)):
            disimg.append(math.sqrt(math.pow((fixcoor[j][0] - 360), 2) + math.pow((fixcoor[j][1] - 288), 2)))
        mdisimg = np.mean(disimg)
        sddisimg = np.var(disimg)

        if fixa>0:
            meanda=dura/fixa
        if fixb>0:
            meandb = durb / fixb
        if fixc>0:
            meandc = durc / fixc
        if fixd>0:
            meandd = durd / fixd
        if fixe>0:
            meande = dure / fixe

        m_duration = round(np.mean(record),4)
        sd_duration = round(np.var(record),4)
        duration=round(duration/mediatime,4)
        dura=round(dura/mediatime,4)
        durb = round(durb / mediatime, 4)
        durc = round(durc / mediatime, 4)
        durd = round(durd / mediatime, 4)
        dure = round(dure / mediatime, 4)


        for data in myRecords_templist1:
            gduration += float(data[5])
            if data[9] == 'A':
                gdura += float(data[5])
            elif data[9] == 'B':
                gdurb += float(data[5])
            elif data[9] == 'C':
                gdurc += float(data[5])
            elif data[9] == 'D':
                gdurd += float(data[5])
            elif data[9] == 'E':
                gdure += float(data[5])
        gduration=round(gduration,4)
        gdura=round(gdura,4)
        gdurb = round(gdurb, 4)
        gdurc = round(gdurc, 4)
        gdurd = round(gdurd, 4)
        gdure = round(gdure, 4)

        features.append([sample.split('.')[0],duration,m_duration,sd_duration,dura,durb,durc,durd,
                         dure,mdisimg,sddisimg,meanda,meandb,meandc,meandd,meande,gduration,gdura,gdurb,gdurc,gdurd,gdure])
    f = xlwt.Workbook()
    sheet3 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    line1=['sample','duration','m_duration','sd_duration','dura','durb','durc','durd',
           'dure','mdisimg','sddisimg','meanda','meandb','meandc','meandd','meande','gduration','gdura','gdurb','gdurc','gdurd','gdure']
    for i in range(len(line1)):
        sheet3.write(0,i,line1[i])

    for i in range(len(features)):
        for j in range(len(features[i])):
            sheet3.write(i+1,j,features[i][j])
    f.save(f'{paramas.rateaoifeatures}/media{file}.xls')
'''