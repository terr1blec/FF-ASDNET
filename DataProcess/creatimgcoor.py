import numpy as np
import os
import pandas as pd
import openpyxl
from multiprocessing import Process
import xlwt
import xlrd

##生成每个样本有效的视频帧索引，相应的坐标，以便后续直接调用，生成显著性拼接图像

class paramas():
    def __init__(self):
        self.datafile = '20210623'
        self.savefile = '1datadecolumn'

        self.savesaccadecoor = '2saccadecoor'
        self.savegazepointcoor = '2gazepointcoor'

        self.savesaccadefixation = '3saccadefixation'
        self.savefixation = '3fixation'
        self.savegazepoint = '3gazepoint'
        self.videostart = 'videostart'

        self.oksaccadefixation = '5oksaccadefixation'
        self.okfixation = '5okfixation'
        self.okgazepoint = '5okgazepoint'
        self.videostartsplit = 'videostartsplit'

        self.fixationsample='6fixationsample'
        self.saccadefixationsample='6saccadefixationsample'
        self.gazepointsample='6gazepointsample'

        self.fixationimgcoor='fixationimg/fixationcoor'
        self.saccadefixationimgcoor = 'saccadefixationimg/fixationcoor'
        self.gazepointimgcoor = 'gazepointimg/fixationcoor'
        self.gazepointfps='gazepointfps'

#视频开始播放的时间
def getVideoStart(path):
    Participants = {}
    participantsList1 = os.listdir(path)
    participantsList = []
    for p in participantsList1:
        participantsList.append(p.split('.')[0])
    for file in participantsList:
        sheet = pd.read_excel(path+'/' + file + '.xlsx', header=None)
        fileList = sheet.values.tolist()
        Participants[file] = fileList  # 一个人一个字典
    return Participants

#获取所有被试三轮实验的数据 存在list放入字典
def getScanpath(path):
    Participants = {}
    participantsList1 = os.listdir(path)
    participantsList = []
    for p in participantsList1:
        participantsList.append(p.split('.')[0])
    for file in participantsList:
        sheet = pd.read_excel(path + '/' + file + '.xlsx',header=None)
        fileList = sheet.values.tolist()
        Participants[file] = fileList  # 一个人一个字典
    return Participants

#Saccadefixation的情况也适用，改变存储位置即可
def Fixationcoor(paramas,scanpaths,video,m):
    fps = {'1': 234, '2': 286, '3': 44, '4': 225, '5': 186, '6': 64, '7': 80, '8': 83, '9': 152, '10': 128, '11': 135,
           '12': 58, '13': 75, '14': 86}

    keys=list(scanpaths.keys())
    for i in range(len(keys)):
        for j in range(len(video[keys[i]])):
            if video[keys[i]][j][2]=='媒体'+m+'.avi':
                T0=video[keys[i]][j][0]
                break

        add = []
        for j in range(len(scanpaths[keys[i]][0])):
            add.append('nan')
        scanpaths[keys[i]].append(add)

        f = xlwt.Workbook()
        sheet3 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
        sheet3.write(0,0,'fps(从0开始）')
        sheet3.write(0, 1, 'coordinateX')
        sheet3.write(0, 2, 'coordinateY')
        sheet3.write(0, 3, 'Duration')
        sheet3.write(0, 4, 'EyemovementType')
        v = 0
        cc=0
        for j in range(1,len(scanpaths[keys[i]])):
            if scanpaths[keys[i]][j][2]=='媒体'+m+'.avi':
                if str(scanpaths[keys[i]][j][7])!='nan':
                    duration=int(scanpaths[keys[i]][j][5])
                    frame=(int(scanpaths[keys[i]][j][0])-T0)//40000  #后续就可以直接忽略0帧和最后的帧
                    length=round(duration/40)
                    for k in range(frame,frame+length+1):
                        if k <= fps[m]:
                            v+=1
                            sheet3.write(v, 0, k)
                            sheet3.write(v, 1, round(float(scanpaths[keys[i]][j][7])*720))
                            sheet3.write(v, 2, round(float(scanpaths[keys[i]][j][8])*576))
                            sheet3.write(v, 3, scanpaths[keys[i]][j][5])
                            sheet3.write(v, 4, scanpaths[keys[i]][j][3])

                        else:
                            cc+=1
        if cc>=4:
            print(f'media{m} {keys[i]} 超出4帧及以上')
        if v>0 and cc<4:
            f.save(paramas.fixationimgcoor+'/media'+m+'/'+keys[i]+'.xls')

def gazepointfps(paramas,scanpaths,video,m):
    fps = {'1': 234, '2': 286, '3': 44, '4': 225, '5': 186, '6': 64, '7': 80, '8': 83, '9': 152, '10': 128, '11': 135,
           '12': 58, '13': 75, '14': 86}

    keys = list(scanpaths.keys())
    for i in range(len(keys)):
        frame=-1
        for j in range(len(video[keys[i]])):
            if video[keys[i]][j][2] == '媒体' + m + '.avi':
                T0 = video[keys[i]][j][0]
                break

        scanpaths[keys[i]][0].append('fps(从0开始)')

        f = openpyxl.Workbook()  # 打开一个将写的文件
        sheet3 = f.create_sheet(index=0)  # 在将写的文件创建sheet
        for k in range(1,len(scanpaths[keys[i]][0])+1):
            sheet3.cell(1,k).value=scanpaths[keys[i]][0][k-1]
        v=1

        for j in range(1,len(scanpaths[keys[i]])):
            scanpaths[keys[i]][j][2]=round(float(scanpaths[keys[i]][j][2])*720)
            scanpaths[keys[i]][j][3] = round(float(scanpaths[keys[i]][j][3]) * 576)
            if scanpaths[keys[i]][j][4]=='媒体'+m+'.avi':
                v+=1
                frame=(int(scanpaths[keys[i]][j][0])-T0)//40000
                scanpaths[keys[i]][j].append(frame)
                for k in range(len(scanpaths[keys[i]][j])):
                    sheet3.cell(v, k+1).value=scanpaths[keys[i]][j][k]

        if frame-fps[m]>=4:
            print(f'media{m} {keys[i]} 超出4帧及以上')
        elif v>1 and frame-fps[m]<4:
            f.save(paramas.gazepointfps+'/media'+m+'/'+keys[i]+'.xlsx')

def gazepointcoor(paramas,m):
    files=os.listdir(paramas.gazepointfps+'/media'+m)
    for file in files:
        sheet=pd.read_excel(paramas.gazepointfps+'/media'+m+'/'+file,usecols=[2,3,4,10])
        sheet=sheet.values.tolist()
        add=[]
        for i in range(len(sheet[0])):
            add.append('nan')
        sheet.append(add)

        f = xlwt.Workbook()
        sheet3 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
        sheet3.write(0,0,'fps')
        sheet3.write(0,1,'coordinateX')
        sheet3.write(0, 2, 'coordinateY')

        fps=sheet[0][3]
        xc,yc=[],[]
        xc.append(sheet[0][0])
        yc.append(sheet[0][1])
        v=0
        for j in range(1,len(sheet)):
            if sheet[j][3]==fps:
                xc.append(sheet[j][0])
                yc.append(sheet[j][1])
            else:
                v+=1
                mxc=np.mean(xc)
                myc=np.mean(yc)
                sheet3.write(v, 0, fps)
                sheet3.write(v, 1, round(mxc))
                sheet3.write(v, 2, round(myc))
                fps=sheet[j][3]
                xc, yc = [], []
                xc.append(sheet[j][0])
                yc.append(sheet[j][1])

        f.save(paramas.gazepointimgcoor+'/media'+m+'/'+file.split('.')[0]+'.xls')


# def run(i):
#     for j in range(i,int(i*13+13)):
#         confps(pa,files[j])

pa=paramas()

videos=getVideoStart(pa.videostartsplit)
scanpath=getScanpath(pa.fixationsample)


for i in range(1, 15):
    if not os.path.exists(pa.fixationimgcoor+'/media'+str(i)):
        os.makedirs(pa.fixationimgcoor+'/media'+str(i))
    if not os.path.exists(pa.saccadefixationimgcoor+'/media'+str(i)):
        os.makedirs(pa.saccadefixationimgcoor+'/media'+str(i))
    if not os.path.exists(pa.gazepointimgcoor+'/media'+str(i)):
        os.makedirs(pa.gazepointimgcoor+'/media'+str(i))
    if not os.path.exists(pa.gazepointfps+'/media'+str(i)):
        os.makedirs(pa.gazepointfps+'/media'+str(i))

# process_list = []
# for i in range(16,20):  #开启5个子进程执行fun1函数
#     p = Process(target=run,args=(i,)) #实例化进程对象
#     p.start()
#     process_list.append(p)
#
# for i in process_list:
#     i.join()

# confps(pa,files[260])

process_list = []
for i in range(1,15):  #开启5个子进程执行fun1函数
    p = Process(target=Fixationcoor,args=(pa,scanpath,videos,str(i))) #实例化进程对象
    p.start()
    process_list.append(p)

for i in process_list:
    i.join()

# files=os.listdir(pa.gazepointsample)



