import numpy as np
import os
import pandas as pd
import cv2
import openpyxl
from multiprocessing import Process
import xlwt
import math
import shutil

"""
绘制动力学图像

执行dvmm(paramas,0.5,视频编号)函数

"""


class paramas:
    def __init__(self):
        self.datafile = "20210623"
        self.savefile = "1datadecolumn"

        self.savesaccadecoor = "2saccadecoor"
        self.savegazepointcoor = "2gazepointcoor"

        self.savesaccadefixation = "3saccadefixation"
        self.savefixation = "3fixation"
        self.savegazepoint = "3gazepoint"
        self.videostart = "videostart"

        self.oksaccadefixation = "5oksaccadefixation"
        self.okfixation = "5okfixation"
        self.okgazepoint = "5okgazepoint"
        self.videostartsplit = "videostartsplit"

        self.fixationsample = "6fixationsample"
        self.saccadefixationsample = "6saccadefixationsample"
        self.gazepointsample = "6gazepointsample"

        self.fixationimgcoor = "fixationimg/fixationcoor"
        self.saccadefixationimgcoor = "saccadefixationimg/fixationcoor"
        self.gazepointimgcoor = "gazepointimg/fixationcoor"
        self.gazepointfps = "gazepointfps"

        self.fixationfsjoint = "fixationimg/fsjoint14ss"
        self.saccadefixationfsjoint = "saccadefixationimg/fsjoint10ss"
        self.gazepointfsjoint = "gazepointimg/fsjoint14ss"
        self.newfixationfsjoint = "newfixationimg/fsjoint10ss"
        self.newsaccadefixationfsjoint = "newsaccadefixationimg/fsjoint10ss"
        self.newgazepointfsjoint = "newgazepointimg/fsjoint10ss"
        self.saliency = "../Saliency"
        self.newsaliency = "../TASED-Net-master/output"
        self.frame = "../img/frame"
        self.gazepointfsjoint1d = "gazepointimg/fsjoint1d"
        self.newgazepointfsjoint1d = "newgazepointimg/fsjoint1d"
        self.fixationfsjoint1d = "fixationimg/fsjoint1d"
        self.newfixationfsjoint1d = "newfixationimg/fsjoint1d"
        self.saccadefixationfsjoint1d = "saccadefixationimg/fsjoint1d"
        self.newsaccadefixationfsjoint1d = "newsaccadefixationimg/fsjoint1d"

        self.saveDynamicVisualization = "DynamicalVisualize/imgwos4_4"
        self.dvdata = "DynamicalVisualize/data"
        self.dvselect = "DynamicalVisualize/img"
        self.delecrudegazepointimgcoor = "gazepointimg/delecrudegazepoint"
        self.root_sample = "../sample/20220817wos4/14"


##计算得的用来进行标准化的值
dv_normalize = {
    0: {
        1: [
            66.15619113267242,
            7.8665293687419595,
            0.9383242157740411,
            0.0,
            0.0,
            1.344811431034462e-16,
        ],
        2: [
            59.43639948511905,
            5.747581849968286,
            0.6891540440205097,
            0.0,
            0.0,
            3.1558441601368873e-10,
        ],
        3: [
            59.34249553495631,
            6.914349328237942,
            0.8059305165976423,
            0.006911447084234205,
            8.327005764919271e-16,
            4.6110794423240457e-17,
        ],
        4: [
            76.02511399535675,
            8.076920166184633,
            0.9686178625337724,
            0.0,
            0.0,
            4.597814124379494e-09,
        ],
        5: [
            76.3837789014031,
            9.065354796897918,
            1.0696178989539975,
            0.0,
            0.0,
            9.633019419005506e-10,
        ],
        6: [
            53.056694913860895,
            6.366294086136416,
            0.7638941788020658,
            0.0,
            0.0,
            7.926384657288761e-08,
        ],
        7: [
            58.62675280374918,
            5.779333909373168,
            0.6775007523885889,
            0.0,
            0.0,
            9.241867330617978e-17,
        ],
        8: [
            74.08370574530433,
            8.629560542917487,
            1.0124219689949363,
            0.0,
            3.9969627671612504e-17,
            4.328806674964921e-08,
        ],
        9: [
            76.42634785946846,
            9.006222804752685,
            1.0758016025812485,
            0.0,
            0.0,
            1.048439858578565e-08,
        ],
        10: [
            64.05143588684996,
            7.614886833236622,
            0.9136927085009438,
            0.0,
            0.0,
            3.293962493876323e-08,
        ],
        11: [
            64.71561972039228,
            7.714801535750444,
            0.91924502183168,
            0.0,
            0.0,
            1.6654011529838542e-18,
        ],
        12: [
            62.5855424455104,
            7.458676180230031,
            0.8889023694239591,
            0.0,
            8.159486472077134e-16,
            6.095080904289468e-08,
        ],
        13: [
            57.216723270550155,
            6.83302286752265,
            0.8180082988434308,
            0.006911447084230795,
            0.0,
            2.8598210142570058e-08,
        ],
        14: [
            71.40696407984376,
            7.530568896928116,
            0.8811571245502298,
            0.0,
            0.0,
            1.070855355643277e-08,
        ],
    },
    1: {
        1: [
            66.15619113267242,
            7.8665293687419595,
            0.9383242157740411,
            0.0,
            0.0,
            1.344811431034462e-16,
        ],
        2: [
            59.43639948511905,
            5.747581849968286,
            0.6891540440205097,
            0.0,
            0.0,
            9.284611427884988e-17,
        ],
        3: [
            59.34249553495631,
            6.914349328237942,
            0.8059305165976423,
            0.0,
            1.7986332452225624e-16,
            1.1354290918400181e-08,
        ],
        4: [
            76.02511399535675,
            8.076920166184633,
            0.9686178625337724,
            0.0,
            0.0,
            4.597814124379494e-09,
        ],
        5: [
            76.3837789014031,
            9.065354796897918,
            1.0696178989539975,
            0.0,
            0.0,
            9.633019419005506e-10,
        ],
        6: [
            53.056694913860895,
            6.366294086136416,
            0.7638941788020658,
            0.0,
            0.0,
            4.55318982317352e-08,
        ],
        7: [
            58.62675280374918,
            6.132037549095701,
            0.7357856430400409,
            0.0,
            0.0,
            9.241867330617978e-17,
        ],
        8: [
            74.08370574530433,
            8.629560542917487,
            1.0124219689949363,
            0.0,
            2.6646418447741666e-16,
            4.328806674964921e-08,
        ],
        9: [
            76.42634785946846,
            9.006222804752685,
            1.0758016025812485,
            0.0,
            0.0,
            1.949969735817614e-08,
        ],
        10: [
            64.05143588684996,
            7.614886833236622,
            0.9136927085009438,
            0.0,
            0.0,
            3.293962493876323e-08,
        ],
        11: [
            64.71561972039228,
            7.714801535750444,
            0.91924502183168,
            0.0,
            0.0,
            1.6654011529838542e-18,
        ],
        12: [
            45.20842414697454,
            4.929822247758651,
            0.5586076149202679,
            0.0,
            8.159486472077134e-16,
            1.7204655136401383e-07,
        ],
        13: [
            57.216723270550155,
            6.83302286752265,
            0.8180082988434308,
            0.006911447084230795,
            4.3966590438773754e-16,
            3.82407828760135e-08,
        ],
        14: [
            71.40696407984376,
            7.530568896928116,
            0.8811571245502298,
            0.0,
            0.0,
            1.070855355643277e-08,
        ],
    },
    2: {
        1: [
            61.70721900374532,
            7.2018280554771446,
            0.8376038054391376,
            0.0,
            0.0,
            1.344811431034462e-16,
        ],
        2: [
            47.80893557018558,
            5.724732731745709,
            0.6865178722374728,
            0.0,
            0.0,
            9.284611427884988e-17,
        ],
        3: [
            52.967992606325616,
            6.273736500890607,
            0.7517579654124776,
            0.0,
            1.7986332452225624e-16,
            4.6110794423240457e-17,
        ],
        4: [
            71.45598339985439,
            6.927465912158238,
            0.8281825302854863,
            0.0,
            0.0,
            4.597814124379494e-09,
        ],
        5: [
            76.3837789014031,
            9.065354796897918,
            1.0696178989539975,
            0.0,
            0.0,
            9.633019419005506e-10,
        ],
        6: [
            53.056694913860895,
            6.366294086136416,
            0.7638941788020658,
            0.0,
            0.0,
            4.55318982317352e-08,
        ],
        7: [
            58.62675280374918,
            6.132037549095701,
            0.7357856430400409,
            0.0,
            0.0,
            9.241867330617978e-17,
        ],
        8: [
            74.08370574530433,
            8.629560542917487,
            1.0124219689949363,
            0.0,
            3.9969627671612504e-17,
            5.142009376836255e-08,
        ],
        9: [
            75.40112450468654,
            7.05085658066876,
            0.8453309617300995,
            0.0,
            0.0,
            1.048439858578565e-08,
        ],
        10: [
            64.05143588684996,
            7.614886833236622,
            0.9136927085009438,
            0.0,
            0.0,
            3.293962493876323e-08,
        ],
        11: [
            64.71561972039228,
            7.714801535750444,
            0.91924502183168,
            0.0,
            0.0,
            1.6654011529838542e-18,
        ],
        12: [
            62.5855424455104,
            7.458676180230031,
            0.8889023694239591,
            0.0,
            4.229636434481405e-07,
            6.095080904289468e-08,
        ],
        13: [
            52.51980651579888,
            6.301872632085298,
            0.7561642227124188,
            0.006911447084230795,
            0.0,
            2.8598210142570058e-08,
        ],
        14: [
            56.41937511568781,
            6.544764782982067,
            0.7789577855163312,
            0.0,
            0.0,
            4.5656606274423676e-08,
        ],
    },
    3: {
        1: [
            66.15619113267242,
            7.8665293687419595,
            0.9383242157740411,
            0.0,
            0.0,
            1.1053128896784429e-09,
        ],
        2: [
            59.43639948511905,
            5.747581849968286,
            0.6891540440205097,
            0.0,
            0.0,
            9.284611427884988e-17,
        ],
        3: [
            59.34249553495631,
            6.914349328237942,
            0.8059305165976423,
            0.0,
            1.7986332452225624e-16,
            4.6110794423240457e-17,
        ],
        4: [
            76.02511399535675,
            8.076920166184633,
            0.9686178625337724,
            0.0,
            0.0,
            4.597814124379494e-09,
        ],
        5: [
            59.52401726106201,
            7.079530803514855,
            0.8468411958807663,
            0.0,
            0.0,
            9.633019419005506e-10,
        ],
        6: [
            42.68904470941953,
            4.3279787766350335,
            0.5123341765047474,
            0.0,
            8.060541580441854e-16,
            4.55318982317352e-08,
        ],
        7: [
            58.62675280374918,
            6.132037549095701,
            0.7357856430400409,
            0.0,
            0.0,
            4.63433553738961e-08,
        ],
        8: [
            57.777951710248786,
            6.871956939914218,
            0.8103682112692457,
            0.0,
            3.9969627671612504e-17,
            4.328806674964921e-08,
        ],
        9: [
            76.42634785946846,
            9.006222804752685,
            1.0758016025812485,
            0.0,
            0.0,
            1.048439858578565e-08,
        ],
        10: [
            52.9930830415292,
            6.359424341957182,
            0.7631614474927615,
            0.0,
            0.0,
            3.293962493876323e-08,
        ],
        11: [
            59.87763132634315,
            6.930950284550986,
            0.8113654244038696,
            0.0,
            0.0,
            5.321566589071185e-08,
        ],
        12: [
            62.5855424455104,
            7.458676180230031,
            0.8889023694239591,
            0.0,
            8.159486472077134e-16,
            6.095080904289468e-08,
        ],
        13: [
            57.216723270550155,
            6.83302286752265,
            0.8180082988434308,
            0.006911447084230795,
            0.0,
            2.8598210142570058e-08,
        ],
        14: [
            71.40696407984376,
            7.530568896928116,
            0.8811571245502298,
            0.0,
            0.0,
            1.070855355643277e-08,
        ],
    },
    4: {
        1: [
            66.15619113267242,
            7.8665293687419595,
            0.9383242157740411,
            0.0,
            0.0,
            1.344811431034462e-16,
        ],
        2: [
            59.43639948511905,
            5.747581849968286,
            0.6891540440205097,
            0.0,
            0.0,
            9.284611427884988e-17,
        ],
        3: [
            59.34249553495631,
            6.914349328237942,
            0.8059305165976423,
            0.0,
            1.7986332452225624e-16,
            4.6110794423240457e-17,
        ],
        4: [
            76.02511399535675,
            8.076920166184633,
            0.9686178625337724,
            0.0,
            0.0,
            4.651251659377919e-09,
        ],
        5: [
            76.3837789014031,
            9.065354796897918,
            1.0696178989539975,
            0.0,
            0.0,
            2.0887827307187148e-09,
        ],
        6: [
            53.056694913860895,
            6.366294086136416,
            0.7638941788020658,
            0.006911447084230795,
            0.0,
            4.55318982317352e-08,
        ],
        7: [
            51.10440093416357,
            6.132037549095701,
            0.7357856430400409,
            0.0,
            8.127157626561209e-16,
            9.241867330617978e-17,
        ],
        8: [
            74.08370574530433,
            8.629560542917487,
            1.0124219689949363,
            0.0,
            3.9969627671612504e-17,
            4.328806674964921e-08,
        ],
        9: [
            76.42634785946846,
            9.006222804752685,
            1.0758016025812485,
            0.0,
            0.0,
            1.048439858578565e-08,
        ],
        10: [
            64.05143588684996,
            7.614886833236622,
            0.9136927085009438,
            0.0,
            0.0,
            3.426547013434961e-08,
        ],
        11: [
            64.71561972039228,
            7.714801535750444,
            0.91924502183168,
            0.0,
            7.860693442083792e-16,
            1.6654011529838542e-18,
        ],
        12: [
            62.5855424455104,
            7.458676180230031,
            0.8889023694239591,
            0.0,
            8.159486472077134e-16,
            6.095080904289468e-08,
        ],
        13: [
            57.216723270550155,
            6.83302286752265,
            0.8180082988434308,
            0.006911447084230795,
            0.0,
            2.8598210142570058e-08,
        ],
        14: [
            71.40696407984376,
            7.530568896928116,
            0.8811571245502298,
            0.0,
            0.0,
            1.070855355643277e-08,
        ],
    },
}


def maxminnormalize(paramas, sca, rou, media):
    fo = open(f"{paramas.root_sample}/{rou}/trainvalidationsample.txt", "r")
    participants = []
    myFile = fo.read()
    myRecords = myFile.split("\n")
    maxv, maxa, maxj = 0, 0, 0
    minv, mina, minj = np.inf, np.inf, np.inf
    for y in range(0, len(myRecords) - 1):
        participants.append(myRecords[y].split("\t"))  # 把表格数据存入
    for i in range(len(participants)):
        if participants[i][0] == "media" + str(media):
            sheet = pd.read_excel(
                f"{paramas.delecrudegazepointimgcoor}{sca}/media{media}/{participants[i][1]}.xlsx"
            )
            sheet = sheet.values.tolist()
            v0, a0 = 0, 0
            if str(sheet[0][2]) == "nan":
                x0 = 360
                y0 = 288
            else:
                x0 = float(sheet[0][2]) * 720
                y0 = float(sheet[0][3]) * 576
            t0 = int(sheet[0][0])

            for k in range(1, len(sheet)):
                if str(sheet[k][3]) != "nan":
                    x = float(sheet[k][2]) * 720
                    y = float(sheet[k][3]) * 576
                    d = math.sqrt(pow(x - x0, 2) + pow(y - y0, 2))
                    v = d / (int(sheet[k][0]) - t0) * 1000

                    a = abs(v - v0) / (int(sheet[k][0]) - t0) * 1000

                    j = abs(a - a0) / (int(sheet[k][0]) - t0) * 1000

                    v0 = v
                    a0 = a
                    x0 = x
                    y0 = y
                    t0 = int(sheet[k][0])

                    if v > maxv:
                        maxv = v
                    if a > maxa:
                        maxa = a
                    if j > maxj:
                        maxj = j
                    if v < minv:
                        minv = v
                    if a < mina:
                        mina = a
                    if j < minj:
                        minj = j
    print(f"{media}:[{maxv},{maxa},{maxj},{minv},{mina},{minj}],")
    return maxv, maxa, maxj, minv, mina, minj


def dvmm(paramas, sca, media):
    for r in range(5):
        [maxv, maxa, maxj, minv, mina, minj] = maxminnormalize(paramas, sca, r, media)
        if not os.path.exists(f"{paramas.saveDynamicVisualization}/{r}/media{media}"):
            os.makedirs(f"{paramas.saveDynamicVisualization}/{r}/media{media}")

        samples = os.listdir(f"{paramas.delecrudegazepointimgcoor}{sca}/media{media}")
        for sample in samples:
            sheet = pd.read_excel(
                f"{paramas.delecrudegazepointimgcoor}{sca}/media{media}/{sample}"
            )
            sheet = sheet.values.tolist()
            img = np.zeros((576, 720, 3))
            v0, a0 = 0, 0
            if str(sheet[0][3]) == "nan":
                x0 = 360
                y0 = 288
            else:
                x0 = float(sheet[0][2]) * 720
                y0 = float(sheet[0][3]) * 576
            t0 = int(sheet[0][0])

            for i in range(1, len(sheet)):
                if str(sheet[i][3]) != "nan":
                    x = float(sheet[i][2]) * 720
                    y = float(sheet[i][3]) * 576
                    d = math.sqrt(pow(x - x0, 2) + pow(y - y0, 2))
                    v = d / (int(sheet[i][0]) - t0) * 1000
                    R = round((v - minv) / (maxv - minv) * 255)
                    a = abs(v - v0) / (int(sheet[i][0]) - t0) * 1000
                    G = round((a - mina) / (maxa - mina) * 255)
                    j = abs(a - a0) / (int(sheet[i][0]) - t0) * 1000
                    B = round((j - minj) / (maxj - minj) * 255)
                    if R > 255:
                        R = 255
                    if R < 0:
                        R = 0
                    if G > 255:
                        G = 255
                    if G < 0:
                        G = 0
                    if B > 255:
                        B = 255
                    if B < 0:
                        B = 0
                    cv2.line(
                        img,
                        (round(x0), round(y0)),
                        (round(x), round(y)),
                        color=(B, G, R),
                        thickness=4,
                        lineType=cv2.LINE_AA,
                    )
                    v0 = v
                    a0 = a
                    x0 = x
                    y0 = y
                    t0 = int(sheet[i][0])
            cv2.imwrite(
                f"{paramas.saveDynamicVisualization}/{r}/media{media}/{sample[:-5]}.jpg",
                img,
            )


# allScanpaths=creatimgcoor.getScanpath('5samples')
# pa=paramas()
# for i in range(1,15):
#     DynamicVisualization(pa,allScanpaths,str(i))
# DynamicVisualization(pa,allScanpaths,str(1))

# fo = open('../sample/participant.txt', 'r')
# participants = []
# myFile = fo.read()
# myRecords = myFile.split('\n')
# for y in range(0, len(myRecords) - 1):
#     participants.append(myRecords[y].split('\t'))  # 把表格数据存入
# medias=os.listdir('newgazepointimg/fsjoint10ss')
# samples={}
# for media in medias:
#     samples[media] = os.listdir('newgazepointimg/fsjoint10ss/' + media)
# for i in range(len(participants)):
#     if not os.path.exists('newgazepointimg/fsjoint10ssp/'+str(i)):
#         os.makedirs('newgazepointimg/fsjoint10ssp/'+str(i))
#     for j in range(len(participants[i])-1):
#         for media in medias:
#             for k in range(len(samples[media])):
#                 if samples[media][k][:len(participants[i][j])]==participants[i][j]:
#                     img=cv2.imread('newgazepointimg/fsjoint10ss/' + media+'/'+samples[media][k])
#                     cv2.imwrite('newgazepointimg/fsjoint10ssp/'+str(i)+'/'+media+'-'+samples[media][k],img)

pa = paramas()
# for i in range(1,14):
#     data(pa,i)
# for i in range(15,34):
#     data(pa,i)
# data(pa,14)


# files=os.listdir(pa.gazepointsample)
# for file in files:
#     DynamicVisualization(pa,file)
process_list = []

for i in range(1, 15):  # 开启5个子进程执行fun1函数
    # if i!=14:
    p = Process(target=dvmm, args=(pa, 0.5, i))  # 实例化进程对象
    p.start()
    process_list.append(p)

for i in process_list:
    i.join()

# medias=os.listdir(pa.gazepointimgcoor)
# for media in medias:
#     files=os.listdir(f'{pa.gazepointimgcoor}/{media}')
#     imgs=os.listdir(f'{pa.saveDynamicVisualization}/{media}')
#     if not os.path.exists(f'{pa.dvselect}/{media}'):
#         os.makedirs(f'{pa.dvselect}/{media}')
#     for file in files:
#         for img in imgs:
#             if file[:-4]==img[:len(file[:-4])]:
#                 shutil.copy(f'{pa.saveDynamicVisualization}/{media}/{img}',f'{pa.dvselect}/{media}/{img}')
#                 break


"""
def data(paramas,file):
    sheet = pd.read_excel(f'{paramas.savefile}/project{file}.xlsx', header=None)
    sheet = sheet.values.tolist()

    if not os.path.exists(paramas.dvdata):
        os.makedirs(paramas.dvdata)
    add = [[]]
    for i in range(len(sheet[0])):
        add[0].append('nan')
    ##删除无用列
    ##de为要删除的列索引
    de = []
    de.extend([1, 2])
    for i in range(4, 23):
        de.append(i)
    for i in range(24, 50):
        de.append(i)
    for i in range(52, 56):
        de.append(i)
    for i in range(60, 64):
        de.append(i)
    sheet = np.delete(sheet, de, 1)
    print(f'file{file} delete column')
    print(len(sheet[0]))

    # 标记样本是第几次实验
    add = [[]]
    for i in range(len(sheet[0])):
        add[0].append('nan')

    sheet=np.append(sheet,add,axis=0)
    participant = sheet[1][1]
    rou = 1
    lenmedia = 0

    de = []
    print(sheet[2][5])
    for i in range(2, len(sheet)):
        if sheet[i][1] == participant:
            if sheet[i][2] == 'VideoStimulusStart':
                lenmedia += 1
                if lenmedia > 14:
                    rou += 1
                    lenmedia = 1
            sheet[i][1] = participant + str(rou)
        else:
            rou = 1
            lenmedia = 0
            participant = sheet[i][1]
            sheet[i][1] = participant + str(rou)
        if not ('媒体' in str(sheet[i][5])):
            de.append(i)
    sheet = np.delete(sheet, de, 0)
    print(f'file{file} delete row (7)')

    add = [[]]
    for i in range(len(sheet[0])):
        add[0].append('nan')
    sheet=np.append(sheet,add,axis=0)
    newsheet = []
    newsheet.append(['Recording timestamp', 'Participant name','Event', 'Gaze point X (MCSnorm)', 'Gaze point Y (MCSnorm)',
                     'Presented Media name','Eye movement type', 'Gaze event duration', 'Eye movement type index'])
    newsheet.append(sheet[1])
    participant = sheet[1][1]

    for i in range(2, len(sheet)):
        if sheet[i][1] == participant:
            newsheet.append(sheet[i])
        else:
            outwb = openpyxl.Workbook()  # 打开一个将写的文件
            outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
            for k in range(1, len(newsheet) + 1):
                for j in range(1, len(newsheet[k - 1]) + 1):
                    outws.cell(k, j).value = newsheet[k - 1][j - 1]  # 写文件
            outwb.save(f'{paramas.dvdata}/{participant}.xlsx')  # 一定要记得保存

            participant = sheet[i][1]
            newsheet = []
            newsheet.append(
                ['Recording timestamp', 'Participant name', 'Event', 'Gaze point X (MCSnorm)', 'Gaze point Y (MCSnorm)',
                 'Presented Media name', 'Eye movement type', 'Gaze event duration', 'Eye movement type index'])
            newsheet.append(sheet[i])

def DynamicVisualization(paramas,file):
    sheet=pd.read_excel(f'{paramas.dvdata}/{file}')
    sheet=sheet.values.tolist()
    add=[]
    for i in range(len(sheet[0])):
        add.append('nan')
    sheet.append(add)
    js=0
    participant=sheet[0][1]
    media=sheet[0][5]
    img = np.zeros((576, 720, 3))
    v0,a0=0,0
    if str(sheet[0][3])=='nan':
        x0=360
        y0=288
    else:
        x0=sheet[0][3]
        y0=sheet[0][4]
    t0=sheet[0][0]

    for i in range(1,len(sheet)):
        if sheet[i][1]==participant:
            if sheet[i][5]==media:
                if str(sheet[i][3])!='nan':
                    js+=1
                    x=sheet[i][3]*720
                    y=sheet[i][4]*576
                    d=math.sqrt(pow(x - x0, 2) + pow(y - y0,2))
                    v=d/(sheet[i][0]-t0)*1000
                    R=round(v/115*8*255)
                    a=abs(v-v0)/(sheet[i][0]-t0)*1000
                    G=round(a/14*8*255)
                    j=abs(a-a0)/(sheet[i][0]-t0)*1000
                    B=round(j/1.8*8*255)
                    cv2.line(img, (round(x0), round(y0)),(round(x), round(y)),color=(B, G, R), thickness=1, lineType=cv2.LINE_AA)
                    v0=v
                    a0=a
                    x0=x
                    y0=y
                    t0=sheet[i][0]
            else:
                if not os.path.exists(f'{paramas.saveDynamicVisualization}/media{media[2:-4]}'):
                    os.makedirs(f'{paramas.saveDynamicVisualization}/media{media[2:-4]}')
                cv2.imencode('.jpg', img)[1].tofile(f'{paramas.saveDynamicVisualization}/media{media[2:-4]}/{participant}.jpg')
                js = 0
                media = sheet[i][5]
                img = np.zeros((576, 720, 3))
                v0, a0 = 0, 0
                if str(sheet[i][3]) == 'nan':
                    x0 = 360
                    y0 = 288
                else:
                    x0 = sheet[i][3]
                    y0 = sheet[i][4]
                t0 = sheet[i][0]
        else:
            if not os.path.exists(f'{paramas.saveDynamicVisualization}/media{media[2:-4]}'):
                os.makedirs(f'{paramas.saveDynamicVisualization}/media{media[2:-4]}')
            cv2.imencode('.jpg', img)[1].tofile(
                f'{paramas.saveDynamicVisualization}/media{media[2:-4]}/{participant}.jpg')
            js = 0
            participant=sheet[i][1]
            media = sheet[i][5]
            img = np.zeros((576, 720, 3))
            v0, a0 = 0, 0
            if str(sheet[i][3]) == 'nan':
                x0 = 360
                y0 = 288
            else:
                x0 = sheet[i][3]
                y0 = sheet[i][4]
            t0 = sheet[i][0]
"""
