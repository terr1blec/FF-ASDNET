import numpy as np
import os
import pandas as pd
import openpyxl
from multiprocessing import Process
import math

##计算扫视坐标

class paramas():
    def __init__(self):
        self.datafile = '20210623'
        self.savefile='1datadecolumn'

        self.savesaccadecoor = '2saccadecoor'
        self.savegazepointcoor = '2gazepointcoor'

def calsaccadecoor(paramas,file):
    sheet=pd.read_excel(f'{paramas.savefile}/project{file}.xlsx',header=None)
    sheet=sheet.values.tolist()
    add=[]
    for i in range(len(sheet[0])):
        add.append('nan')
    sheet.append(add)

    coorx,coory,co=[],[],[]
    nan=-1  #非Saccade
    index=-1
    if sheet[1][57]=='Saccade':
        index=sheet[1][59]
        co.append(1)
        if str(sheet[1][50])!='nan':
            nan=0
            coorx.append(sheet[1][50])
            coory.append(sheet[1][51])
        else:
            nan=1

    for i in range(2,len(sheet)):
        print(i+1)
        if nan==-1:  #上一次记录非Saccade
            coorx, coory, co = [], [], []
            if sheet[i][57]=='Saccade':
                index=sheet[i][59]
                co.append(i)
                if str(sheet[i][50])!='nan':
                    nan=0
                    coorx.append(float(sheet[i][50]))
                    coory.append(float(sheet[i][51]))
                else:
                    nan=1
            else:
                coorx, coory, co = [], [], []
                nan=-1
                index=-1

        elif nan==0:  #上一次记录是Saccade，且坐标不为nan
            if sheet[i][57]=='Saccade' and sheet[i][59]==index:  #与上一次记录属于同一条记录
                co.append(i)
                if str(sheet[i][50])=='nan':
                    print('同一条记录上一个坐标不是nan这一个是')
                    nan=1
                else:
                    coorx.append(float(sheet[i][50]))
                    coory.append(float(sheet[i][51]))
            elif sheet[i][57]=='Saccade' and sheet[i][59]!=index:
                if len(coorx)!=0:
                    newx=np.mean(coorx)
                    newy = np.mean(coory)
                    for j in co:
                        sheet[j][62]=newx
                        sheet[j][63]=newy
                index=sheet[i][59]
                coorx, coory, co = [], [], []
                co.append(i)
                if str(sheet[i][50])=='nan':
                    nan=1
                else:
                    nan=0
                    coorx.append(float(sheet[i][50]))
                    coory.append(float(sheet[i][51]))
            elif sheet[i][57]!='Saccade':
                nan=-1
                index=-1
                if len(coorx)!=0:
                    newx = np.mean(coorx)
                    newy = np.mean(coory)
                    for j in co:
                        sheet[j][62] = newx
                        sheet[j][63] = newy
                coorx, coory, co = [], [], []

        elif nan==1:
            if sheet[i][57]=='Saccade' and sheet[i][59]==index:
                co.append(i)
                if str(sheet[i][50])!='nan':
                    print('同一条记录上一个坐标是nan这一个不是')
                    nan=0
                    coorx.append(float(sheet[i][50]))
                    coory.append(float(sheet[i][51]))
            elif sheet[i][57]=='Saccade' and sheet[i][59]!=index:
                if len(coorx)!=0:
                    newx=np.mean(coorx)
                    newy = np.mean(coory)
                    for j in co:
                        sheet[j][62]=newx
                        sheet[j][63]=newy
                index=sheet[i][59]
                coorx, coory, co = [], [], []
                co.append(i)
                if str(sheet[i][50])=='nan':
                    nan=1
                else:
                    nan=0
                    coorx.append(float(sheet[i][50]))
                    coory.append(float(sheet[i][51]))
            elif sheet[i][57]!='Saccade':
                if len(coorx)!=0:
                    newx=np.mean(coorx)
                    newy = np.mean(coory)
                    for j in co:
                        sheet[j][62]=newx
                        sheet[j][63]=newy
                nan=-1
                coorx, coory, co = [], [], []
                index=-1
    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(sheet) + 1):
        for j in range(1, len(sheet[i - 1]) + 1):
            outws.cell(i, j).value = sheet[i - 1][j - 1]  # 写文件
    outwb.save(f'{paramas.savesaccadecoor}/project{file}.xlsx')  # 一定要记得保存

#计算所有gaze point的坐标，连续三次记录做一个平均
def calgazepointcoor(paramas,file):
    sheet=pd.read_excel(f'{paramas.savefile}/project{file}.xlsx',header=None)
    sheet=sheet.values.tolist()
    add=[]
    for i in range(len(sheet[0])):
        add.append('nan')
    sheet.append(add)

    coorx,coory,co=[],[],[]
    index=0
    nan=1
    if str(sheet[1][50])!='nan':
        nan=0
        index+=1
        co.append(1)
        coorx.append(sheet[1][50])
        coory.append(sheet[1][51])

    for i in range(2,len(sheet)):
        print(i+1)
        if nan==0:
            if str(sheet[i][50])!='nan':
                index+=1
                co.append(i)
                coorx.append(float(sheet[i][50]))
                coory.append(float(sheet[i][51]))
                if index==5:
                    x=np.mean(coorx)
                    y=np.mean(coory)
                    for j in co:
                        sheet[j][50]=x
                        sheet[j][51]=y
                    index=0
                    coorx,coory,co=[],[],[]
            else:
                nan=1
                if index!=0:
                    x = np.mean(coorx)
                    y = np.mean(coory)
                    for j in co:
                        sheet[j][50] = x
                        sheet[j][51] = y
                index=0
                coorx,coory,co=[],[],[]
        else:
            if str(sheet[i][50])!='nan':
                nan=0
                index+=1
                co.append(i)
                coorx.append(float(sheet[i][50]))
                coory.append(float(sheet[i][51]))

    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(sheet) + 1):
        for j in range(1, len(sheet[i - 1]) + 1):
            outws.cell(i, j).value = sheet[i - 1][j - 1]  # 写文件
    outwb.save(f'{paramas.savegazepointcoor}/project{file}.xlsx')  # 一定要记得保存

def run(paramas, r):
    for i in range(int(r * 3 + 1), int(r * 3 + 4)):
        calgazepointcoor(paramas, i)

pa=paramas()
if not os.path.exists(pa.savesaccadecoor):
    os.makedirs(pa.savesaccadecoor)
if not os.path.exists(pa.savegazepointcoor):
    os.makedirs(pa.savegazepointcoor)
process_list = []
for i in range(11):  #开启5个子进程执行fun1函数
    p = Process(target=run,args=(pa,i)) #实例化进程对象
    p.start()
    process_list.append(p)

for i in process_list:
    i.join()

# for i in range(0,5):  #开启5个子进程执行fun1函数
#     p = Process(target=run,args=(pa,i)) #实例化进程对象
#     p.start()
#     process_list.append(p)
#
# for i in process_list:
#     i.join()

# calcoordinate(pa,33)