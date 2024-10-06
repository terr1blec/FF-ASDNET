import numpy as np
import pandas as pd
import os
import xlwt
from multiprocessing import Process
import openpyxl

##删除无用的列数据

class paramas():
    def __init__(self):
        self.datafile='20210623'
        self.savefile='1datadecolumn'

def condense(paramas,file):
    sheet=pd.read_excel(f'{paramas.datafile}/Project{file}（2020.02.01） Data Export.xlsx',header=None)
    sheet=sheet.values.tolist()

    ##删除无用列
    ##de为要删除的列索引
    de=[]
    de.extend([3,4])
    for i in range(6,22):
        de.append(i)
    de.extend([42,75])
    for i in range(77,83):
        de.append(i)
    for i in range(90,165):
        de.append(i)
    sheet=np.delete(sheet,de,1)
    print(f'file{file} delete column')
    print(len(sheet[0]))

    ##删除无用行
    # de=[]
    # for i in range(1,len(sheet)):
    #     if (sheet[i][56] in ['1.jpg','2.jpg','3.jpg','4.jpg','f82b9097df6fe3383cae212be078fe2.jpg','1610688457(1).jpg']):
    #         de.append(i)
    # sheet=np.delete(sheet,de,0)
    # print(f'file{file} delete row (56)')
    #
    # index1,index2,de=[],[],[]
    # for i in range(1,len(sheet)):
    #     if sheet[i][23]=='RecordingStart':
    #         index1.append(i+1)
    #     elif sheet[i][23]=='Eye tracker Calibration end':
    #         index2.append(i+1)
    # for i in range(len(index2)):
    #     for j in range(index1[i],index2[i]):
    #         de.append(j)
    # sheet=np.delete(sheet,de,0)
    # print(f'file{file} delete row (23)')

    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1,len(sheet)+1):
        for j in range(1,len(sheet[i-1])+1):
            outws.cell(i, j).value = sheet[i-1][j-1]  # 写文件
    outwb.save(f'{paramas.savefile}/project{file}.xlsx')  # 一定要记得保存

def run(paramas,r):
    for i in range(int(r*6+1),int(r*6+7)):
        condense(paramas,i)

pa=paramas()
if not os.path.exists(pa.savefile):
    os.makedirs(pa.savefile)
process_list = []
for i in range(31,34):  #开启5个子进程执行fun1函数
    p = Process(target=condense,args=(pa,i)) #实例化进程对象
    p.start()
    process_list.append(p)

for i in process_list:
    i.join()
#
# for i in range(5):  #开启5个子进程执行fun1函数
#     p = Process(target=run,args=(pa,i)) #实例化进程对象
#     p.start()
#     process_list.append(p)
#
# for i in process_list:
#     i.join()