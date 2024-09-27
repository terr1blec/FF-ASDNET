import os
import pandas as pd
import openpyxl

class parameter():
    def __init__(self):
        self.lr_scheduler='plateau'
        self.gamma=0.5
        self.epochs=50
        self.lr=1e-6
        self.batch_size=32
        self.flush_history=0
        self.patience=5
        self.save_model=1
        self.log_every=100
        self.version='v20221204/dv_zssa'

args=parameter()
samplepath = '../sample/20220817wos4'
prefix_name = f'l_4fc_d0.5_convcon_conv512_1*1_conv256_3*3_update_1loss_3fc/d0.5_n0.5_shuffle'
batch = 64
lr = 5e-7
best_epoch = 40
ws = 12

newsheet=[]
row1=['media','name','isASD','x','y']
newsheet.append(row1)
for roun in range(5):
    pathfiles = os.listdir(f'{args.version}/output/{prefix_name}_{roun}/{batch}_{lr}')
    for file in pathfiles:
        if file[10:ws] == str(best_epoch):
            pathfile = f'{args.version}/output/{prefix_name}_{roun}/{batch}_{lr}/{file}'
    fo = open(pathfile,'r')
    val = []
    myFile = fo.read()
    myRecords = myFile.split('\n')
    for y in range(0, len(myRecords) - 1):
        val.append(myRecords[y].split('\t'))

    foo=open(f'../sample/20220817wos4/14/{roun}/testsample.txt','r')
    samples=[]
    myFile = foo.read()
    myRecords = myFile.split('\n')
    for y in range(0, len(myRecords) - 1):
        samples.append(myRecords[y].split('\t'))

    # newsheet=[]
    # row1=['media','name','isASD','x','y']
    # newsheet.append(row1)

    for s in range(len(val)):
        # if val[s][0]==val[s][1]:
        sheet = pd.read_excel(f'../DataProcessing/gazepointimg/delecrudegazepoint0.5/{samples[s][0]}/{samples[s][1]}.xlsx')
        sheet = sheet.values.tolist()
        js,x,y = 0,0,0
        for i in range(len(sheet)):
            if str(sheet[i][2]) != 'nan':
                js+=1
                x += float(sheet[i][2])
                y += float(sheet[i][3])
        x=round((x*720)/js)
        y = round((y * 576) / js)
        newsheet.append([samples[s][0],samples[s][1],val[s][0],x,y])

outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
for i in range(1, len(newsheet) + 1):
    for j in range(1, len(newsheet[i - 1]) + 1):
        outws.cell(i, j).value = newsheet[i - 1][j - 1]  # 写文件
outwb.save(f'all.xlsx')
