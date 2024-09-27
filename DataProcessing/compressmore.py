import numpy as np
import os
import pandas as pd
import openpyxl
from multiprocessing import Process

#进一步删除无用行数据

# 保存文件的路径
class paramas:
    def __init__(self):
        self.datafile = "20210623"
        self.savefile = "1datadecolumn_JY"

        self.savesaccadecoor = "2saccadecoor_JY"
        self.savegazepointcoor = "2gazepointcoor"

        self.savesaccadefixation = "3saccadefixation"
        self.savefixation = "3fixation"
        self.savegazepoint = "3gazepoint"
        self.videostart = "videostart"
        self.videoend = "videoend"
        self.videostartrow = "videostartrow_JY"
        self.videoendrow = "videoendrow_JY"


# 删除更多无用列，除Fixation和Saccade以外的数据，除media播放过程以外的数据
# 保留Fixation和Saccade
def savefixation_saccade(paramas, file):
    sheet = pd.read_excel(f"{paramas.savesaccadecoor}/project{file}.xlsx", header=None)
    sheet = sheet.values.tolist()

    # 标记样本是第几次实验
    add = []
    for i in range(len(sheet[0])):
        add.append("nan")
    sheet.append(add)
    participant = sheet[1][3]
    rou = 1
    lenmedia = 0

    for i in range(2, len(sheet)):
        if sheet[i][3] == participant:
            if sheet[i][23] == "VideoStimulusStart":
                lenmedia += 1
                if lenmedia > 14:
                    rou += 1
                    lenmedia = 1
            sheet[i][3] = participant + str(rou)
        else:
            rou = 1
            lenmedia = 0
            participant = sheet[i][3]

    ##删除无用列
    ##de为要删除的列索引
    de = []
    de.append(2)
    for i in range(4, 22):
        de.append(i)
    for i in range(23, 50):
        de.append(i)
    for i in range(52, 56):
        de.append(i)
    de.extend([60, 61])
    sheet = np.delete(sheet, de, 1)
    print(f"file{file} delete column")
    print(len(sheet[0]))

    ##删除无用行
    de = []
    for i in range(1, len(sheet)):
        if not ("媒体" in sheet[i][6]):
            de.append(i)
    sheet = np.delete(sheet, de, 0)
    print(f"file{file} delete row (6)")

    de = []
    for i in range(1, len(sheet)):
        if not (sheet[i][7] in ["Saccade", "Fixation"]):
            de.append(i)
    sheet = np.delete(sheet, de, 0)
    print(f"file{file} delete row (7)")

    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(sheet) + 1):
        for j in range(1, len(sheet[i - 1]) + 1):
            outws.cell(i, j).value = sheet[i - 1][j - 1]  # 写文件
    outwb.save(f"{paramas.savesaccadefixation}/project{file}.xlsx")  # 一定要记得保存


# 另一种处理方法：保留所有Gaze point
def saveallgazepoint(paramas, file):
    sheet = pd.read_excel(
        f"{paramas.savegazepointcoor}/project{file}.xlsx", header=None
    )
    sheet = sheet.values.tolist()

    # 标记样本是第几次实验
    add = []
    for i in range(len(sheet[0])):
        add.append("nan")
    sheet.append(add)
    participant = sheet[1][3]
    rou = 1
    lenmedia = 0

    for i in range(2, len(sheet)):
        if sheet[i][3] == participant:
            if sheet[i][23] == "VideoStimulusStart":
                lenmedia += 1
                if lenmedia > 14:
                    rou += 1
                    lenmedia = 1
            sheet[i][3] = participant + str(rou)
        else:
            rou = 1
            lenmedia = 0
            participant = sheet[i][3]

    ##删除无用列
    ##de为要删除的列索引
    de = []
    de.append(2)
    for i in range(4, 22):
        de.append(i)
    for i in range(23, 50):
        de.append(i)
    for i in range(52, 56):
        de.append(i)
    de.extend([60, 61])
    sheet = np.delete(sheet, de, 1)
    print(f"file{file} delete column")
    print(len(sheet[0]))

    ##删除无用行
    de = []
    for i in range(1, len(sheet)):
        if not ("媒体" in sheet[i][6]):
            de.append(i)
        else:
            if str(sheet[i][4]) == "nan":
                de.append(i)
    sheet = np.delete(sheet, de, 0)
    print(f"file{file} delete row (4,6)")

    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(sheet) + 1):
        for j in range(1, len(sheet[i - 1]) + 1):
            outws.cell(i, j).value = sheet[i - 1][j - 1]  # 写文件
    outwb.save(f"{paramas.savegazepoint}/project{file}.xlsx")  # 一定要记得保存


# 另一种处理方法：只保留fixation
def savefixation(paramas, file):
    sheet = pd.read_excel(f"{paramas.savesaccadecoor}/project{file}.xlsx", header=None)
    sheet = sheet.values.tolist()

    # 标记样本是第几次实验
    add = []
    for i in range(len(sheet[0])):
        add.append("nan")
    sheet.append(add)
    participant = sheet[1][3]
    rou = 1
    lenmedia = 0

    for i in range(2, len(sheet)):
        if sheet[i][3] == participant:
            if sheet[i][23] == "VideoStimulusStart":
                lenmedia += 1
                if lenmedia > 14:
                    rou += 1
                    lenmedia = 1
            sheet[i][3] = participant + str(rou)
        else:
            rou = 1
            lenmedia = 0
            participant = sheet[i][3]

    ##删除无用列
    ##de为要删除的列索引
    de = []
    de.append(2)
    for i in range(4, 22):
        de.append(i)
    for i in range(23, 50):
        de.append(i)
    for i in range(52, 56):
        de.append(i)
    de.extend([60, 61])
    sheet = np.delete(sheet, de, 1)
    print(f"file{file} delete column")
    print(len(sheet[0]))

    ##删除无用行
    de = []
    for i in range(1, len(sheet)):
        if not ("媒体" in sheet[i][6]):
            de.append(i)
        else:
            if str(sheet[i][7]) != "Fixation":
                de.append(i)
    sheet = np.delete(sheet, de, 0)
    print(f"file{file} delete row (7)")

    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(sheet) + 1):
        for j in range(1, len(sheet[i - 1]) + 1):
            outws.cell(i, j).value = sheet[i - 1][j - 1]  # 写文件
    outwb.save(f"{paramas.savefixation}/project{file}.xlsx")  # 一定要记得保存


def videostart(paramas, file):
    sheet = pd.read_excel(f"{paramas.savesaccadecoor}/{file}", header=None)
    sheet = sheet.values.tolist()

    # 标记样本是第几次实验
    add = []
    for i in range(len(sheet[0])):
        add.append("nan")
    sheet.append(add)

    newsheet = []
    time = 0
    media = 0

    for i in range(2, len(sheet)):
        if sheet[i][23] == "VideoStimulusStart":
            for j in range(i + 1, len(sheet)):
                if sheet[j][23] == "TTL out":
                    time = j
                    break
            for j in range(time + 1, len(sheet)):
                if str(sheet[j][56]) != "nan":
                    media = j
                    break
            newsheet.append([time, sheet[time][3], sheet[media][56]])

    lenmedia = 0
    rou = 1
    participant = newsheet[0][1]
    for i in range(len(newsheet)):
        if newsheet[i][1] == participant:
            lenmedia += 1
            if lenmedia > 14:
                rou += 1
                lenmedia = 1
        else:
            participant = newsheet[i][1]
            rou = 1
            lenmedia = 1
        newsheet[i][1] = newsheet[i][1] + str(rou)
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(newsheet) + 1):
        for j in range(1, len(newsheet[i - 1]) + 1):
            outws.cell(i, j).value = newsheet[i - 1][j - 1]  # 写文件
    outwb.save(f"{paramas.videostartrow}/{file}")  # 一定要记得保存


def videoend(paramas, file):
    sheet = pd.read_excel(f"{paramas.savesaccadecoor}/{file}", header=None)
    sheet = sheet.values.tolist()

    # 标记样本是第几次实验
    add = []
    for i in range(len(sheet[0])):
        add.append("nan")
    sheet.append(add)

    newsheet = []
    time = 0
    media = 0

    for i in range(2, len(sheet)):
        if sheet[i][23] == "VideoStimulusEnd":
            time = i
            for j in range(1, time):
                if str(sheet[time - j][56]) != "nan":
                    media = time - j
                    break
            newsheet.append([time, sheet[time][3], sheet[media][56]])

    lenmedia = 0
    rou = 1
    participant = newsheet[0][1]
    for i in range(len(newsheet)):
        if newsheet[i][1] == participant:
            lenmedia += 1
            if lenmedia > 14:
                rou += 1
                lenmedia = 1
        else:
            participant = newsheet[i][1]
            rou = 1
            lenmedia = 1
        newsheet[i][1] = newsheet[i][1] + str(rou)
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(newsheet) + 1):
        for j in range(1, len(newsheet[i - 1]) + 1):
            outws.cell(i, j).value = newsheet[i - 1][j - 1]  # 写文件
    outwb.save(f"{paramas.videoendrow}/{file}")  # 一定要记得保存


def test(paramas, file):
    sheet = pd.read_excel(f"{paramas.savesaccadecoor}/project{file}.xlsx", header=None)
    sheet = sheet.values.tolist()

    # 标记样本是第几次实验
    add = []
    for i in range(len(sheet[0])):
        add.append("nan")
    sheet.append(add)

    newsheet = []
    time = 0
    media = 0

    for i in range(2, len(sheet)):
        if sheet[i][23] == "VideoStimulusStart":
            for j in range(i + 1, len(sheet)):
                if sheet[j][23] == "TTL out":
                    time = j
                    break
            for j in range(time + 1, len(sheet)):
                if str(sheet[j][56]) != "nan":
                    media = j
                    break
            newsheet.append([sheet[time][0], sheet[time][3], sheet[media][56]])

    lenmedia = 0
    rou = 1
    participant = newsheet[0][1]
    print(participant + str(rou))
    for i in range(len(newsheet)):
        if newsheet[i][1] == participant:
            lenmedia += 1
            if lenmedia > 14:
                rou += 1
                lenmedia = 1
                print(participant + str(rou))
        else:
            participant = newsheet[i][1]
            rou = 1
            lenmedia = 1
        newsheet[i][1] = newsheet[i][1] + str(rou)


def run(paramas, r):
    for i in range(int(r * 3 + 1), int(r * 3 + 4)):
        test(paramas, i)


pa = paramas()
if not os.path.exists(pa.savesaccadefixation):
    os.makedirs(pa.savesaccadefixation)
if not os.path.exists(pa.savegazepoint):
    os.makedirs(pa.savegazepoint)
if not os.path.exists(pa.savefixation):
    os.makedirs(pa.savefixation)
if not os.path.exists(pa.videostart):
    os.makedirs(pa.videostart)
if not os.path.exists(pa.videostartrow):
    os.makedirs(pa.videostartrow)
if not os.path.exists(pa.videoendrow):
    os.makedirs(pa.videoendrow)

process_list = []
files = os.listdir(f"{pa.savefile}")
for i in range(len(files)):  # 开启5个子进程执行fun1函数
    p = Process(target=videoend, args=(pa, files[i]))  # 实例化进程对象
    p.start()
    process_list.append(p)

for i in process_list:
    i.join()
