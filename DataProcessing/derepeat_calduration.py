import numpy as np
import os
import pandas as pd
import openpyxl
from multiprocessing import Process

##删除连续重复，计算duration

class paramas():
    def __init__(self):
        self.datafile = '20210623'
        self.savefile = '1datadecolumn'

        self.savesaccadecoor = '2saccadecoor'
        self.savegazepointcoor = '2gazepointcoor'

        self.savesaccadefixation = '3saccadefixation'
        self.savefixation = '3fixation'
        self.savegazepoint = '3gazepoint'
        self.videostart = 'videostart'

        self.oksaccadefixation='5oksaccadefixation'
        self.okfixation='5okfixation'
        self.okgazepoint='5okgazepoint'
        self.videostartsplit='videostartsplit'

#fixation的改变存储位置即可
def saccadefixationderepeat(paramas,file):
    sheet = pd.read_excel(f'{paramas.savefixation}/project{file}.xlsx', header=None)
    sheet = sheet.values.tolist()
    add=[]
    for i in range(len(sheet[0])):
        add.append(-1)
    sheet.append(add)
    newsheet=[]
    newsheet.append(['Recording timestamp','Participant name','Presented Media name','Eye movement type','Gaze event duration',
                     'Duration','Eye movement type index','Fixation point X (MCSnorm)','Fixation point Y (MCSnorm)'])

    participant=sheet[1][2]
    media=sheet[1][6]
    eyetype=sheet[1][7]
    typeindex=sheet[1][9]
    timer=[]
    timer.append(sheet[1][0])
    gazeduration=sheet[1][8]
    x=sheet[1][10]
    y=sheet[1][11]

    for i in range(2,len(sheet)):
        if sheet[i][2]==participant:
            if sheet[i][6]==media and sheet[i][7]==eyetype and sheet[i][9]==typeindex:
                timer.append(sheet[i][0])
            else:
                if len(timer)>1:
                    duration=round((float(timer[len(timer)-1])-float(timer[0]))/1000)+8
                else:
                    duration=gazeduration
                newsheet.append([timer[0],participant,media,eyetype,gazeduration,duration,typeindex,x,y])
                media = sheet[i][6]
                eyetype = sheet[i][7]
                typeindex = sheet[i][9]
                timer = []
                timer.append(sheet[i][0])
                gazeduration = sheet[i][8]
                x = sheet[i][10]
                y = sheet[i][11]
        else:
            ##保存文件
            outwb = openpyxl.Workbook()  # 打开一个将写的文件
            outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
            for k in range(1, len(newsheet) + 1):
                for j in range(1, len(newsheet[k - 1]) + 1):
                    outws.cell(k, j).value = newsheet[k - 1][j - 1]  # 写文件
            outwb.save(f'{paramas.okfixation}/{participant}.xlsx')  # 一定要记得保存

            media = sheet[i][6]
            eyetype = sheet[i][7]
            typeindex = sheet[i][9]
            timer = []
            timer.append(sheet[i][0])
            gazeduration = sheet[i][8]
            x = sheet[i][10]
            y = sheet[i][11]
            participant=sheet[i][2]
            newsheet=[]
            newsheet.append(['Recording timestamp', 'Participant name', 'Presented Media name', 'Eye movement type',
                             'Gaze event duration',
                             'Duration', 'Eye movement type index', 'Fixation point X (MCSnorm)',
                             'Fixation point Y (MCSnorm)'])

def gazepointderepeat(paramas,file):
    sheet = pd.read_excel(f'{paramas.savegazepoint}/project{file}.xlsx', header=None)
    sheet = sheet.values.tolist()
    add=[]
    for i in range(len(sheet[0])):
        add.append(-1)
    sheet.append(add)
    newsheet=[]
    newsheet.append(['Recording timestamp','Participant name','Gaze point X (MCSnorm)','Gaze point Y (MCSnorm)','Presented Media name',
                     'Eye movement type','Gaze event duration','Eye movement type index','Fixation point X (MCSnorm)',
                     'Fixation point Y (MCSnorm)'])
    newsheet.append([sheet[1][0],sheet[1][2],sheet[1][4],sheet[1][5],sheet[1][6],sheet[1][7],sheet[1][8],sheet[1][9],sheet[1][10]
                     ,sheet[1][11]])

    participant=sheet[1][2]

    for i in range(2,len(sheet)):
        if sheet[i][2]==participant:
            newsheet.append(
                [sheet[i][0], sheet[i][2], sheet[i][4], sheet[i][5], sheet[i][6], sheet[i][7], sheet[i][8], sheet[i][9],
                 sheet[i][10], sheet[i][11]])
        else:
            outwb = openpyxl.Workbook()  # 打开一个将写的文件
            outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
            for k in range(1, len(newsheet) + 1):
                for j in range(1, len(newsheet[k - 1]) + 1):
                    outws.cell(k, j).value = newsheet[k - 1][j - 1]  # 写文件
            outwb.save(f'{paramas.okgazepoint}/{participant}.xlsx')  # 一定要记得保存

            participant=sheet[i][2]
            newsheet=[]
            newsheet.append(
                ['Recording timestamp', 'Participant name', 'Gaze point X (MCSnorm)', 'Gaze point Y (MCSnorm)',
                 'Presented Media name',
                 'Eye movement type', 'Gaze event duration', 'Eye movement type index', 'Fixation point X (MCSnorm)',
                 'Fixation point Y (MCSnorm)'])
            newsheet.append(
                [sheet[i][0], sheet[i][2], sheet[i][4], sheet[i][5], sheet[i][6], sheet[i][7], sheet[i][8], sheet[i][9],
                 sheet[i][10], sheet[i][11]])

def videosplit(paramas,file):
    sheet = pd.read_excel(f'{paramas.videostart}/project{file}.xlsx', header=None)
    sheet = sheet.values.tolist()
    add = []
    for i in range(len(sheet[0])):
        add.append(-1)
    sheet.append(add)
    newsheet=[]
    newsheet.append(sheet[0])
    participant=sheet[0][1]
    for i in range(1,len(sheet)):
        if sheet[i][1]==participant:
            newsheet.append(sheet[i])
        else:
            outwb = openpyxl.Workbook()  # 打开一个将写的文件
            outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
            for k in range(1, len(newsheet) + 1):
                for j in range(1, len(newsheet[k - 1]) + 1):
                    outws.cell(k, j).value = newsheet[k - 1][j - 1]  # 写文件
            outwb.save(f'{paramas.videostartsplit}/{participant}.xlsx')  # 一定要记得保存
            newsheet=[]
            participant=sheet[i][1]
            newsheet.append(sheet[i])


def run(paramas, r):
    for i in range(int(r * 3 + 1), int(r * 3 + 4)):
        videosplit(paramas, i)

pa=paramas()
if not os.path.exists(pa.oksaccadefixation):
    os.makedirs(pa.oksaccadefixation)
if not os.path.exists(pa.okgazepoint):
    os.makedirs(pa.okgazepoint)
if not os.path.exists(pa.okfixation):
    os.makedirs(pa.okfixation)
if not os.path.exists(pa.videostartsplit):
    os.makedirs(pa.videostartsplit)

# process_list = []
# for i in range(11):  #开启5个子进程执行fun1函数
#     p = Process(target=run,args=(pa,i)) #实例化进程对象
#     p.start()
#     process_list.append(p)
#
# for i in process_list:
#     i.join()

# saccadefixationderepeat(pa,14)
videosplit(pa,14)
# gazepointderepeat(pa,14)

