import numpy as np
import pandas as pd
import os
import xlwt
from multiprocessing import Process
import openpyxl

"""
数据介绍：原始眼动数据包含很多无用信息，眼动仪频率是120Hz，所以差不多8-9ms记录一次眼动，原始xlsx文件里每两行之间的记录时间戳的差值差不多为8-9ms，而视频一帧是40ms。
为了后续分析,需要将眼动坐标定位到每一帧上。

眼动仪自动将眼动行为分为四种类型类型，包括Fixation，Saccade，Unclassified，EyesnotFound，除了EyesnotFound没有Gaze point坐标，其他眼动类型都有。
在眼动仪中是根据眼球移动的角速度来分类这三种眼动行为的， 连续发生的一种眼动行为用一个eye movement index表示。

研究分为两部分：眼动模式分析以及建立ASD分类模型。在眼动模式分析部分仅使用眼动仪判定为Fixation的数据，并利用这些数据建立基线模型。后续建立ASD分类模型使用所有的Gaze point数据。
程序从原始数据中提取Fixation数据与Gaze point数据并保存。

按顺序执行程序中的相应函数，执行以下工作：
1.删除多余列  condense(paramas,原始数据名)
2。删除除视频播放时间段以外的所有行。每个人观看了3轮，在人名处加上序号以表明当前是第几轮观看视频。
2.1.保留所有Gaze point数据   savecrudeallgazepoint(paramas,1处理后的数据名)
2.2.只保留Fixation数据   savefixation(paramas,1处理后的数据名)
2.2.1.眼动仪将连续发生的一种眼动行为用一个eye movement index表示，连续发生的Fixation行为仅作为一个Fixation，只有一个Fixation坐标，因此删除多余重复的Fixation数据，
排除视频开始播放前以及视频播放完后的注视时间，并将注视时间调整得更加精确。将所有人观看每一轮视频的数据切分成独立的表格并进行保存。
   saccadefixationderepeat(paramas,2.2处理后的数据名)
2.3.保存被试观看时，每个视频开始播放时的时间   videostart(paramas,1处理后的数据名)
2.4.保存被试观看时，每个视频开始播放时的时间   videoend(paramas,1处理后的数据名)
2.5.将视频播放文件进行切分   videosplit(paramas,2.3或2.4处理后的数据名)
3.将Gaze point数据中观看每一轮视频的数据切分成独立的表格并进行保存   crudegazepointsplit(paramas,2.1处理后的数据名)
4.将Gaze point观看每一轮视频的数据进一步切分成每个视频独立的表格并进行保存，得到所有样本   crudegazepointmediasplit(paramas,视频编号)
5.删除Gaze point样本中眼动记录过少的样本   deletecrudegazepoint(paramas,无效眼动的比例,视频编号)，相应的，2.2.1中的Fixation样本中也应删除这些样本
6.保存视频每一帧上的眼动坐标
6.1.Gaze point样本中
6.1.1计算当前眼动时刻对应的视频帧数   gazepointfps(paramas,5处理后的所有样本,2.4处理后的视频播放时间,视频编号)   
6.1.2.计算坐标   gazepointcoor(paramas,视频编号)
6.2.Fixation样本中，直接保存坐标   Fixationcoor(paramas,2.2.1处理后的所有样本,2.4处理后的视频播放时间,视频编号)
"""


"""
各种路径，存放数据
"""


class paramas:
    def __init__(self):
        self.datafile = "../data/data_first/数据20210623"  # 原始数据存放的文件路径
        self.savefile = (
            "1datadecolumn"  # 原始数据删除列后保存的文件路径（1处理后保存的路径）
        )

        self.savesaccadecoor = "2saccadecoor_JY"
        self.savegazepointcoor = "2gazepointcoor"

        self.savesaccadefixation = "3saccadefixation"
        self.savefixation = "3fixation"  # 进一步删除无用列，以及除视频播放时间段以外的所有眼动数据，将表格保存在这里（只有Fixation数据）（2.2处理后保存的路径）
        self.savegazepoint = "3gazepoint"
        self.videostart = "videostart"  # 保存被试观看时，每个视频开始播放的时间（2.3处理后保存的路径）
        self.crudegazepoint = "3crudegazepoint"  # 进一步删除无用列，以及除视频播放时间段以外的所有眼动数据，将表格保存在这里（所有Gaze point数据）（2.1处理后保存的路径）

        self.oksaccadefixation = "5oksaccadefixation"
        self.okfixation = "5okfixation"  # 将Fixation删除多余重复，且将注视时长精确化后，将表格保存在这里（2.2.1处理后保存的路径）
        self.okgazepoint = "5okgazepoint"
        self.okcrudegazepoint = "5okcrudegazepoint"  # 将Gaze point数据中观看每一轮视频的数据切分成独立的表格并进行保存（3处理后保存的路径）
        self.videostartsplit = "videostartsplit"  # 将视频播放时间数据切分后保存的路径（2.5处理后保存的路径）
        self.videoendsplit = (
            "videoendsplit"  # 将视频播放时间数据切分后保存的路径（2.5处理后保存的路径）
        )
        self.videoend = (
            "videoend"  # 保存被试观看时，每个视频结束播放的时间（2.4处理后保存的路径）
        )
        self.videoendrow = "videoendrow"
        self.videostartrow = "videostartrow"
        self.videostartrowsplit = "videostartrowsplit"
        self.videoendrowsplit = "videoendrowsplit"

        # 处理完的一些样本是不能用的，能用的样本在这里
        self.fixationsample = "6fixationsample"  # 2.2.1处理完并保存之后对应/mnt/shareEx/huangmin/ASDeyetracking/data/data_first/2021.7.11.xlsx，选择入组的样本数据，保存在其中
        self.saccadefixationsample = "6saccadefixationsample"
        self.gazepointsample = "6gazepointsample"
        self.crudegazepointsample = "6crudegazepointsample"  # 3处理完并保存之后对应/mnt/shareEx/huangmin/ASDeyetracking/data/data_first/2021.7.11.xlsx，选择入组的样本数据，保存在其中

        self.fixationimgcoor = "fixationimg/fixationcoor"  # 保存眼动坐标（Fixation样本）（6.2处理后保存的路径）
        self.saccadefixationimgcoor = "saccadefixationimg/fixationcoor"
        self.gazepointimgcoor = "gazepointimg/fixationcoor"  # 计算并保存眼动坐标（Gaze point样本）（6.1.2处理后保存的路径）
        self.crudegazepointimgcoor = (
            "gazepointimg/crudegazepoint"  # 每个视频每一轮分开保存（4处理后保存的路径）
        )
        self.delecrudegazepointimgcoor = "gazepointimg/delecrudegazepoint"  # 删除Gaze point样本中眼动记录过少的样本保存的路径（5处理后保存的路径）
        self.missingsamples = "gazepointimg/missingsamples"
        self.xy = "gazepointimg/indexxy"
        self.gazepointfps = "gazepointfps"  # 计算Gaze point样本中当前眼动时刻对应的视频帧数保存的路径（6。1.1处理后保存的路径）


"""
1.删除多余列   condense(paramas,原始数据名)
删除原始文件中无用的列，新文件保存在paramas.savefile = '1datadecolumn'
保留了Recording timestamp	Computer timestamp	Sensor	Participant name	Average calibration accuracy (mm)
Average calibration precision SD (mm)	Average calibration precision RMS (mm)	Average calibration accuracy (degrees)
Average calibration precision SD (degrees)	Average calibration precision RMS (degrees)	Average calibration accuracy (pixels)
Average calibration precision SD (pixels)	Average calibration precision RMS (pixels)	Average validation accuracy (mm)
Average validation precision SD (mm)	Average validation precision RMS (mm)	Average validation accuracy (degrees)
Average validation precision SD (degrees)	Average validation precision RMS (degrees)	Average validation accuracy (pixels)
Average validation precision SD (pixels)	Average validation precision RMS (pixels)	Eyetracker timestamp	Event
Gaze point X	Gaze point Y	Gaze point left X	Gaze point left Y	Gaze point right X	Gaze point right Y
Gaze direction left X	Gaze direction left Y	Gaze direction left Z	Gaze direction right X	Gaze direction right Y
Gaze direction right Z	Pupil diameter left	Pupil diameter right	Validity left	Validity right	Eye position left X (DACSmm)
Eye position left Y (DACSmm)	Eye position left Z (DACSmm)	Eye position right X (DACSmm)	Eye position right Y (DACSmm)
Eye position right Z (DACSmm)	Gaze point left X (DACSmm)	Gaze point left Y (DACSmm)	Gaze point right X (DACSmm)
Gaze point right Y (DACSmm)	Gaze point X (MCSnorm)	Gaze point Y (MCSnorm)	Gaze point left X (MCSnorm)	Gaze point left Y (MCSnorm)
Gaze point right X (MCSnorm)	Gaze point right Y (MCSnorm)	Presented Media name	Eye movement type	Gaze event duration
Eye movement type index	Fixation point X	Fixation point Y	Fixation point X (MCSnorm)	Fixation point Y (MCSnorm)

最后有用的记录是Recording timestamp（每一条记录的时间戳），Participant name，Gaze point X (MCSnorm)	Gaze point Y (MCSnorm)（眼睛位置投射在视频上的坐标归一化），
Presented Media name（当前播放的视频），Eye movement type，Gaze event duration	Eye movement type index Fixation point X	Fixation point Y	
Fixation point X (MCSnorm)	Fixation point Y (MCSnorm)
"""


def condense(paramas, file):
    sheet = pd.read_excel(f"{paramas.datafile}/{file}", header=None)
    sheet = sheet.values.tolist()

    ##删除无用列
    # de为要删除的列索引
    de = []
    de.extend([3, 4])
    for i in range(6, 22):
        de.append(i)
    de.extend([42, 75])
    for i in range(77, 83):
        de.append(i)
    for i in range(90, 165):
        de.append(i)
    sheet = np.delete(sheet, de, 1)
    print(f"file{file} delete column")
    print(len(sheet[0]))

    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(sheet) + 1):
        for j in range(1, len(sheet[i - 1]) + 1):
            outws.cell(i, j).value = sheet[i - 1][j - 1]  # 写文件
    outwb.save(f"{paramas.savefile}/{file}")  # 保存


"""
2。1删除除视频播放时间段以外的所有行
保留所有Gaze point数据   savecrudeallgazepoint(paramas,1处理后的数据名)
处理后的数据保存在paramas.crudegazepoint='3crudegazepoint'
"""


def savecrudeallgazepoint(paramas, file):
    sheet = pd.read_excel(f"{paramas.savefile}/{file}", header=None)
    sheet = sheet.values.tolist()

    # 标记样本是第几次实验
    add = []
    for i in range(len(sheet[0])):
        add.append("nan")
    sheet.append(add)
    participant = sheet[1][3]
    rou = 1
    lenmedia = 0

    for i in range(2, len(sheet)):
        if sheet[i][3] == participant:
            if sheet[i][23] == "VideoStimulusStart":
                lenmedia += 1
                if lenmedia > 14:
                    rou += 1
                    lenmedia = 1
            sheet[i][3] = participant + str(rou)
        else:
            rou = 1
            lenmedia = 0
            participant = sheet[i][3]

    ##删除无用列
    # de为要删除的列索引
    de = []
    de.append(2)
    for i in range(4, 22):
        de.append(i)
    for i in range(23, 50):
        de.append(i)
    for i in range(52, 56):
        de.append(i)
    de.extend([60, 61])
    sheet = np.delete(sheet, de, 1)
    print(f"file{file} delete column")
    print(len(sheet[0]))

    ##删除无用行
    # de为要删除的行索引
    de = []
    for i in range(1, len(sheet)):
        if not ("媒体" in sheet[i][6]):
            de.append(i)
        # else:
        #     if str(sheet[i][4])=='nan':
        #         de.append(i)
    sheet = np.delete(sheet, de, 0)
    print(f"file{file} delete row (4,6)")

    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(sheet) + 1):
        for j in range(1, len(sheet[i - 1]) + 1):
            outws.cell(i, j).value = sheet[i - 1][j - 1]  # 写文件
    outwb.save(f"{paramas.crudegazepoint}/{file}")  # 一定要记得保存


"""
2。2删除除视频播放时间段以外的所有行
只保留Fixation数据   savefixation(paramas,1处理后的数据名)
处理后的数据保存在paramas.savefixation = '3fixation'
"""


def savefixation(paramas, file):
    sheet = pd.read_excel(f"{paramas.savefile}/project{file}.xlsx", header=None)
    sheet = sheet.values.tolist()

    # 标记样本是第几次实验
    add = []
    for i in range(len(sheet[0])):
        add.append("nan")
    sheet.append(add)
    participant = sheet[1][3]
    rou = 1
    lenmedia = 0

    for i in range(2, len(sheet)):
        if sheet[i][3] == participant:
            if sheet[i][23] == "VideoStimulusStart":
                lenmedia += 1
                if lenmedia > 14:
                    rou += 1
                    lenmedia = 1
            sheet[i][3] = participant + str(rou)
        else:
            rou = 1
            lenmedia = 0
            participant = sheet[i][3]

    ##删除无用列
    ##de为要删除的列索引
    de = []
    de.append(2)
    for i in range(4, 22):
        de.append(i)
    for i in range(23, 50):
        de.append(i)
    for i in range(52, 56):
        de.append(i)
    de.extend([60, 61])
    sheet = np.delete(sheet, de, 1)
    print(f"file{file} delete column")
    print(len(sheet[0]))

    ##删除无用行
    de = []
    for i in range(1, len(sheet)):
        if not ("媒体" in sheet[i][6]):
            de.append(i)
        else:
            if str(sheet[i][7]) != "Fixation":
                de.append(i)
    sheet = np.delete(sheet, de, 0)
    print(f"file{file} delete row (7)")

    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(sheet) + 1):
        for j in range(1, len(sheet[i - 1]) + 1):
            outws.cell(i, j).value = sheet[i - 1][j - 1]  # 写文件
    outwb.save(f"{paramas.savefixation}/project{file}.xlsx")  # 一定要记得保存


"""
2.2.1.眼动仪将连续发生的一种眼动行为用一个eye movement index表示，连续发生的Fixation行为仅作为一个Fixation，只有一个Fixation坐标，因此删除多余重复的Fixation数据，
并排除视频开始播放前以及视频播放完后的注视时间，从而将注视时间调整得更加精确。保存数据时将每个样本的每一轮数据分开保存。
   saccadefixationderepeat(paramas,2.2处理后的数据名)
处理后的数据保存在paramas.okfixation = '5okfixation'
"""


def saccadefixationderepeat(paramas, file):
    sheet = pd.read_excel(f"{paramas.savefixation}/project{file}.xlsx", header=None)
    sheet = sheet.values.tolist()
    add = []
    for i in range(len(sheet[0])):
        add.append(-1)
    sheet.append(add)
    newsheet = []
    newsheet.append(
        [
            "Recording timestamp",
            "Participant name",
            "Presented Media name",
            "Eye movement type",
            "Gaze event duration",
            "Duration",
            "Eye movement type index",
            "Fixation point X (MCSnorm)",
            "Fixation point Y (MCSnorm)",
        ]
    )

    participant = sheet[1][2]
    media = sheet[1][6]
    eyetype = sheet[1][7]
    typeindex = sheet[1][9]
    timer = []
    timer.append(sheet[1][0])
    gazeduration = sheet[1][8]
    x = sheet[1][10]
    y = sheet[1][11]

    for i in range(2, len(sheet)):
        if sheet[i][2] == participant:
            if (
                sheet[i][6] == media
                and sheet[i][7] == eyetype
                and sheet[i][9] == typeindex
            ):
                timer.append(sheet[i][0])
            else:
                if len(timer) > 1:
                    duration = (
                        round((float(timer[len(timer) - 1]) - float(timer[0])) / 1000)
                        + 8
                    )
                else:
                    duration = gazeduration
                newsheet.append(
                    [
                        timer[0],
                        participant,
                        media,
                        eyetype,
                        gazeduration,
                        duration,
                        typeindex,
                        x,
                        y,
                    ]
                )
                media = sheet[i][6]
                eyetype = sheet[i][7]
                typeindex = sheet[i][9]
                timer = []
                timer.append(sheet[i][0])
                gazeduration = sheet[i][8]
                x = sheet[i][10]
                y = sheet[i][11]
        else:
            ##保存文件
            outwb = openpyxl.Workbook()  # 打开一个将写的文件
            outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
            for k in range(1, len(newsheet) + 1):
                for j in range(1, len(newsheet[k - 1]) + 1):
                    outws.cell(k, j).value = newsheet[k - 1][j - 1]  # 写文件
            outwb.save(f"{paramas.okfixation}/{participant}.xlsx")  # 一定要记得保存

            media = sheet[i][6]
            eyetype = sheet[i][7]
            typeindex = sheet[i][9]
            timer = []
            timer.append(sheet[i][0])
            gazeduration = sheet[i][8]
            x = sheet[i][10]
            y = sheet[i][11]
            participant = sheet[i][2]
            newsheet = []
            newsheet.append(
                [
                    "Recording timestamp",
                    "Participant name",
                    "Presented Media name",
                    "Eye movement type",
                    "Gaze event duration",
                    "Duration",
                    "Eye movement type index",
                    "Fixation point X (MCSnorm)",
                    "Fixation point Y (MCSnorm)",
                ]
            )


"""
处理完之后对应/mnt/shareEx/huangmin/ASDeyetracking/data/data_first/2021.7.11.xlsx，选择入组的样本数据，保存在paramas.fixationsample='6fixationsample'
"""

"""
3.将Gaze point数据中观看每一轮视频的数据切分成独立的表格并进行保存
   crudegazepointsplit(paramas,2.1处理后的数据名)
处理后的数据保存在paramas.okcrudegazepoint='5okcrudegazepoint
"""


def crudegazepointsplit(paramas, file):
    sheet = pd.read_excel(f"{paramas.crudegazepoint}/{file}", header=None)
    sheet = sheet.values.tolist()
    add = []
    for i in range(len(sheet[0])):
        add.append(-1)
    sheet.append(add)
    newsheet = []
    newsheet.append(
        [
            "Recording timestamp",
            "Participant name",
            "Gaze point X (MCSnorm)",
            "Gaze point Y (MCSnorm)",
            "Presented Media name",
            "Eye movement type",
            "Gaze event duration",
            "Eye movement type index",
            "Fixation point X (MCSnorm)",
            "Fixation point Y (MCSnorm)",
        ]
    )
    newsheet.append(
        [
            sheet[1][0],
            sheet[1][2],
            sheet[1][4],
            sheet[1][5],
            sheet[1][6],
            sheet[1][7],
            sheet[1][8],
            sheet[1][9],
            sheet[1][10],
            sheet[1][11],
        ]
    )

    participant = sheet[1][2]

    for i in range(2, len(sheet)):
        if sheet[i][2] == participant:
            newsheet.append(
                [
                    sheet[i][0],
                    sheet[i][2],
                    sheet[i][4],
                    sheet[i][5],
                    sheet[i][6],
                    sheet[i][7],
                    sheet[i][8],
                    sheet[i][9],
                    sheet[i][10],
                    sheet[i][11],
                ]
            )
        else:
            outwb = openpyxl.Workbook()  # 打开一个将写的文件
            outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
            for k in range(1, len(newsheet) + 1):
                for j in range(1, len(newsheet[k - 1]) + 1):
                    outws.cell(k, j).value = newsheet[k - 1][j - 1]  # 写文件
            outwb.save(
                f"{paramas.okcrudegazepoint}/{participant}.xlsx"
            )  # 一定要记得保存

            participant = sheet[i][2]
            newsheet = []
            newsheet.append(
                [
                    "Recording timestamp",
                    "Participant name",
                    "Gaze point X (MCSnorm)",
                    "Gaze point Y (MCSnorm)",
                    "Presented Media name",
                    "Eye movement type",
                    "Gaze event duration",
                    "Eye movement type index",
                    "Fixation point X (MCSnorm)",
                    "Fixation point Y (MCSnorm)",
                ]
            )
            newsheet.append(
                [
                    sheet[i][0],
                    sheet[i][2],
                    sheet[i][4],
                    sheet[i][5],
                    sheet[i][6],
                    sheet[i][7],
                    sheet[i][8],
                    sheet[i][9],
                    sheet[i][10],
                    sheet[i][11],
                ]
            )


"""
处理完之后对应/mnt/shareEx/huangmin/ASDeyetracking/data/data_first/2021.7.11.xlsx，选择入组的样本数据，保存在paramas.crudegazepointsample='6crudegazepointsample'
"""

"""
4.将Gaze point观看每一轮视频的数据进一步切分成每个视频独立的表格并进行保存 
  crudegazepointmediasplit(paramas,视频编号)
处理后的数据保存在paramas.crudegazepointimgcoor='gazepointimg/crudegazepoint'
"""


def crudegazepointmediasplit(paramas, file):
    """
    将原始的眼动数据按照媒体文件进行分割并保存。

    参数：
    paramas (object): 包含参数的对象。
    file (int): 媒体文件的编号。

    返回：
    无返回值。
    """
    if not os.path.exists(f"{paramas.crudegazepointimgcoor}/media{file}"):
        os.makedirs(f"{paramas.crudegazepointimgcoor}/media{file}")

    samples = os.listdir(f"{paramas.crudegazepointsample}")
    for sample in samples:
        sheet = pd.read_excel(f"{paramas.crudegazepointsample}/{sample}", header=None)
        sheet = sheet.values.tolist()
        add = []
        for i in range(len(sheet[0])):
            add.append(-1)
        sheet.append(add)
        newsheet = []
        newsheet.append(sheet[0])

        for i in range(1, len(sheet)):
            if sheet[i][4] == "媒体" + str(file) + ".avi":
                newsheet.append(sheet[i])
        outwb = openpyxl.Workbook()  # 打开一个将写的文件
        outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
        for k in range(1, len(newsheet) + 1):
            for j in range(1, len(newsheet[k - 1]) + 1):
                outws.cell(k, j).value = newsheet[k - 1][j - 1]  # 写文件
        outwb.save(
            f"{paramas.crudegazepointimgcoor}/media{file}/{sample}"
        )  # 一定要记得保存


"""
5.删除Gaze point样本中眼动记录过少的样本
   deletecrudegazepoint(paramas,无效眼动的比例,视频编号)
处理后的数据保存在paramas.delecrudegazepointimgcoor = 'gazepointimg/delecrudegazepoint'
"""


def deletecrudegazepoint(paramas, sca, media):
    if not os.path.exists(f"{paramas.delecrudegazepointimgcoor}{sca}/media{media}"):
        os.makedirs(f"{paramas.delecrudegazepointimgcoor}{sca}/media{media}")
    files = os.listdir(f"{paramas.crudegazepointimgcoor}/media{media}")
    # print(f'media{media}共有{len(files)}个样本')
    dele = 0
    for file in files:
        sheet = pd.read_excel(
            f"{paramas.crudegazepointimgcoor}/media{media}/{file[:-5]}.xlsx",
            header=None,
        )
        sheet = sheet.values.tolist()
        nan = 0
        for i in range(1, len(sheet)):
            if str(sheet[i][2]) == "nan":
                nan += 1
        if nan / (len(sheet) - 1) < sca:
            outwb = openpyxl.Workbook()  # 打开一个将写的文件
            outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
            for k in range(1, len(sheet) + 1):
                for j in range(1, len(sheet[k - 1]) + 1):
                    outws.cell(k, j).value = sheet[k - 1][j - 1]  # 写文件
            outwb.save(
                f"{paramas.delecrudegazepointimgcoor}{sca}/media{media}/{file[:-4]}.xlsx"
            )  # 一定要记得保存
        else:
            dele += 1
    print(f"media{media}删除了有效样本量小于{sca}的{dele}/{len(files)}样本")


"""
2.3.保存被试观看时，每个视频开始播放时的时间   
   videostart(paramas,1处理后的数据名)
处理后的数据保存在paramas.videostart = 'videostart'
"""


def videostart(paramas, file):
    sheet = pd.read_excel(f"{paramas.savefile}/{file}", header=None)
    sheet = sheet.values.tolist()

    # 标记样本是第几次实验
    add = []
    for i in range(len(sheet[0])):
        add.append("nan")
    sheet.append(add)

    newsheet = []
    time = 0
    media = 0

    for i in range(2, len(sheet)):
        if sheet[i][23] == "VideoStimulusStart":
            for j in range(i + 1, len(sheet)):
                if sheet[j][23] == "TTL out":
                    time = j
                    break
            for j in range(time + 1, len(sheet)):
                if str(sheet[j][56]) != "nan":
                    media = j
                    break
            newsheet.append([sheet[time][0], sheet[time][3], sheet[media][56]])

    lenmedia = 0
    rou = 1
    participant = newsheet[0][1]
    for i in range(len(newsheet)):
        if newsheet[i][1] == participant:
            lenmedia += 1
            if lenmedia > 14:
                rou += 1
                lenmedia = 1
        else:
            participant = newsheet[i][1]
            rou = 1
            lenmedia = 1
        newsheet[i][1] = newsheet[i][1] + str(rou)
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(newsheet) + 1):
        for j in range(1, len(newsheet[i - 1]) + 1):
            outws.cell(i, j).value = newsheet[i - 1][j - 1]  # 写文件
    outwb.save(f"{paramas.videostart}/{file}")  # 一定要记得保存


"""
2.4.保存被试观看时，每个视频结束播放时的时间   
   videoend(paramas,1处理后的数据名)
处理后的数据保存在paramas.videoend = 'videoend'
"""


def videoend(paramas, file):
    sheet = pd.read_excel(f"{paramas.savefile}/{file}", header=None)
    sheet = sheet.values.tolist()

    # 标记样本是第几次实验
    add = []
    for i in range(len(sheet[0])):
        add.append("nan")
    sheet.append(add)

    newsheet = []
    time = 0
    media = 0

    for i in range(2, len(sheet)):
        if sheet[i][23] == "VideoStimulusEnd":
            time = i
            for j in range(1, time):
                if str(sheet[time - j][56]) != "nan":
                    media = time - j
                    break
            newsheet.append([time, sheet[time][3], sheet[media][56]])

    lenmedia = 0
    rou = 1
    participant = newsheet[0][1]
    for i in range(len(newsheet)):
        if newsheet[i][1] == participant:
            lenmedia += 1
            if lenmedia > 14:
                rou += 1
                lenmedia = 1
        else:
            participant = newsheet[i][1]
            rou = 1
            lenmedia = 1
        newsheet[i][1] = newsheet[i][1] + str(rou)
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(newsheet) + 1):
        for j in range(1, len(newsheet[i - 1]) + 1):
            outws.cell(i, j).value = newsheet[i - 1][j - 1]  # 写文件
    outwb.save(f"{paramas.videoend}/{file}")  # 一定要记得保存


"""
2.5.将视频播放文件进行切分
   videosplit(paramas,2.3或2.4处理后的数据名)
处理后的数据保存在paramas.videostartsplit = 'videostartsplit' 或paramas.videoendsplit = 'videoendsplit'（要在代码中修改） 
"""


def videosplit(paramas, file):
    if not os.path.exists(f"{paramas.videostartsplit}"):
        os.makedirs(f"{paramas.videostartsplit}")
    sheet = pd.read_excel(f"{paramas.videostart}/{file}", header=None)
    sheet = sheet.values.tolist()
    add = []
    for i in range(len(sheet[0])):
        add.append(-1)
    sheet.append(add)
    newsheet = []
    newsheet.append(sheet[0])
    participant = sheet[0][1]
    for i in range(1, len(sheet)):
        if sheet[i][1] == participant:
            newsheet.append(sheet[i])
        else:
            outwb = openpyxl.Workbook()  # 打开一个将写的文件
            outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
            for k in range(1, len(newsheet) + 1):
                for j in range(1, len(newsheet[k - 1]) + 1):
                    outws.cell(k, j).value = newsheet[k - 1][j - 1]  # 写文件
            outwb.save(
                f"{paramas.videostartsplit}/{participant}.xlsx"
            )  # 一定要记得保存
            newsheet = []
            participant = sheet[i][1]
            newsheet.append(sheet[i])


"""
6.计算当前眼动时刻对应的视频帧数，并保存眼动坐标
"""


# 视频开始播放的时间
def getVideoStart(path):
    Participants = {}
    participantsList1 = os.listdir(path)
    participantsList = []
    for p in participantsList1:
        participantsList.append(p.split(".")[0])
    for file in participantsList:
        sheet = pd.read_excel(path + "/" + file + ".xlsx", header=None)
        fileList = sheet.values.tolist()
        Participants[file] = fileList  # 一个人一个字典
    return Participants


# 获取所有被试三轮实验的数据 存在list放入字典.
# 路径根据所需为Gaze point与Fixation样本进行选择。
def getScanpath(path):
    Participants = {}
    participantsList1 = os.listdir(path)
    participantsList = []
    for p in participantsList1:
        participantsList.append(p.split(".")[0])
    for file in participantsList:
        sheet = pd.read_excel(path + "/" + file + ".xlsx", header=None)
        fileList = sheet.values.tolist()
        Participants[file] = fileList  # 一个人一个字典
    return Participants


"""
6.1.Gaze point样本中
"""
"""
6.1.1.计算当前眼动时刻对应的视频帧数
   gazepointfps(paramas,5处理后的所有样本,2.4处理后的视频播放时间,视频编号)
处理后的数据保存在paramas.gazepointfps='gazepointfps'
"""
"""def gazepointfps(paramas,scanpaths,video,m):
    print(m)
    fps = {'1': 234, '2': 286, '3': 44, '4': 225, '5': 186, '6': 64, '7': 80, '8': 83, '9': 152, '10': 128, '11': 135,
           '12': 58, '13': 75, '14': 86}

    keys = list(scanpaths.keys())
    for i in range(len(keys)):
        frame=-1
        for j in range(len(video[keys[i]])):
            if video[keys[i]][j][2] == '媒体' + m + '.avi':
                T0 = video[keys[i]][j][0]
                break

        scanpaths[keys[i]][0].append('fps(从0开始)')

        f = openpyxl.Workbook()  # 打开一个将写的文件
        sheet3 = f.create_sheet(index=0)  # 在将写的文件创建sheet
        for k in range(1,len(scanpaths[keys[i]][0])+1):
            sheet3.cell(1,k).value=scanpaths[keys[i]][0][k-1]
        v=1

        for j in range(1,len(scanpaths[keys[i]])):
            if str(scanpaths[keys[i]][j][2])!='nan':
                scanpaths[keys[i]][j][2]=round(float(scanpaths[keys[i]][j][2])*720)
                scanpaths[keys[i]][j][3] = round(float(scanpaths[keys[i]][j][3]) * 576)
                if scanpaths[keys[i]][j][4]=='媒体'+m+'.avi':
                    v+=1
                    frame=(int(scanpaths[keys[i]][j][0])-T0)//40000
                    scanpaths[keys[i]][j].append(frame)
                    for k in range(len(scanpaths[keys[i]][j])):
                        sheet3.cell(v, k+1).value=scanpaths[keys[i]][j][k]

        if frame-fps[m]>=4:
            print(f'media{m} {keys[i]} 超出4帧及以上')
        elif v>1 and frame-fps[m]<4:
            f.save(paramas.gazepointfps+'/media'+m+'/'+keys[i]+'.xlsx')
"""


def gazepointfps(paramas, scanpaths, video, m):
    fps = {
        "1": 234,
        "2": 286,
        "3": 44,
        "4": 225,
        "5": 186,
        "6": 64,
        "7": 80,
        "8": 83,
        "9": 152,
        "10": 128,
        "11": 135,
        "12": 58,
        "13": 75,
        "14": 86,
    }

    keys = list(scanpaths.keys())
    for i in range(len(keys)):
        frame = -1
        for j in range(len(video[keys[i]])):
            if video[keys[i]][j][2] == "媒体" + m + ".avi":
                T0 = video[keys[i]][j][0]
                break

        scanpaths[keys[i]][0].append("fps(从0开始)")

        f = openpyxl.Workbook()  # 打开一个将写的文件
        sheet3 = f.create_sheet(index=0)  # 在将写的文件创建sheet
        for k in range(1, len(scanpaths[keys[i]][0]) + 1):
            sheet3.cell(1, k).value = scanpaths[keys[i]][0][k - 1]
        v = 1

        for j in range(1, len(scanpaths[keys[i]])):
            if str(scanpaths[keys[i]][j][2]) != "nan":
                scanpaths[keys[i]][j][2] = round(float(scanpaths[keys[i]][j][2]) * 720)
                scanpaths[keys[i]][j][3] = round(float(scanpaths[keys[i]][j][3]) * 576)
                if scanpaths[keys[i]][j][4] == "媒体" + m + ".avi":
                    v += 1
                    frame = (int(scanpaths[keys[i]][j][0]) - T0) // 40000
                    scanpaths[keys[i]][j].append(frame)
                    for k in range(len(scanpaths[keys[i]][j])):
                        sheet3.cell(v, k + 1).value = scanpaths[keys[i]][j][k]

        if frame - fps[m] >= 4:
            print(f"media{m} {keys[i]} 超出4帧及以上")
        elif v > 1 and frame - fps[m] < 4:
            f.save(paramas.gazepointfps + "/media" + m + "/" + keys[i] + ".xlsx")


"""           
6.1.2.计算当前眼动时刻对应的视频帧数
   gazepointcoor(paramas,视频编号)
处理后的数据保存在paramas.gazepointimgcoor='gazepointimg/fixationcoor'
"""


def gazepointcoor(paramas, m):
    files = os.listdir(paramas.gazepointfps + "/media" + m)
    for file in files:
        sheet = pd.read_excel(
            paramas.gazepointfps + "/media" + m + "/" + file, usecols=[2, 3, 4, 10]
        )
        sheet = sheet.values.tolist()
        add = []
        for i in range(len(sheet[0])):
            add.append("nan")
        sheet.append(add)

        f = xlwt.Workbook()
        sheet3 = f.add_sheet("sheet1", cell_overwrite_ok=True)  # 创建sheet
        sheet3.write(0, 0, "fps")
        sheet3.write(0, 1, "coordinateX")
        sheet3.write(0, 2, "coordinateY")

        fps = sheet[0][3]
        xc, yc = [], []
        xc.append(sheet[0][0])
        yc.append(sheet[0][1])
        v = 0
        for j in range(1, len(sheet)):
            if sheet[j][3] == fps:
                xc.append(sheet[j][0])
                yc.append(sheet[j][1])
            else:
                v += 1
                mxc = np.mean(xc)
                myc = np.mean(yc)
                sheet3.write(v, 0, fps)
                sheet3.write(v, 1, round(mxc))
                sheet3.write(v, 2, round(myc))
                fps = sheet[j][3]
                xc, yc = [], []
                xc.append(sheet[j][0])
                yc.append(sheet[j][1])

        f.save(
            paramas.gazepointimgcoor + "/media" + m + "/" + file.split(".")[0] + ".xls"
        )


"""
6.2.Fixation样本中
   Fixationcoor(paramas,2.2.1处理后的所有样本,2.4处理后的视频播放时间,视频编号)
处理后的数据保存在paramas.fixationimgcoor='fixationimg/fixationcoor'
"""


def Fixationcoor(paramas, scanpaths, video, m):
    fps = {
        "1": 234,
        "2": 286,
        "3": 44,
        "4": 225,
        "5": 186,
        "6": 64,
        "7": 80,
        "8": 83,
        "9": 152,
        "10": 128,
        "11": 135,
        "12": 58,
        "13": 75,
        "14": 86,
    }

    keys = list(scanpaths.keys())
    for i in range(len(keys)):
        for j in range(len(video[keys[i]])):
            if video[keys[i]][j][2] == "媒体" + m + ".avi":
                T0 = video[keys[i]][j][0]
                break

        add = []
        for j in range(len(scanpaths[keys[i]][0])):
            add.append("nan")
        scanpaths[keys[i]].append(add)

        f = xlwt.Workbook()
        sheet3 = f.add_sheet("sheet1", cell_overwrite_ok=True)  # 创建sheet
        sheet3.write(0, 0, "fps(从0开始）")
        sheet3.write(0, 1, "coordinateX")
        sheet3.write(0, 2, "coordinateY")
        sheet3.write(0, 3, "Duration")
        sheet3.write(0, 4, "EyemovementType")
        v = 0
        cc = 0
        for j in range(1, len(scanpaths[keys[i]])):
            if scanpaths[keys[i]][j][2] == "媒体" + m + ".avi":
                if str(scanpaths[keys[i]][j][7]) != "nan":
                    duration = int(scanpaths[keys[i]][j][5])
                    frame = (
                        int(scanpaths[keys[i]][j][0]) - T0
                    ) // 40000  # 后续就可以直接忽略0帧和最后的帧
                    length = round(duration / 40)
                    for k in range(frame, frame + length + 1):
                        if k <= fps[m]:
                            v += 1
                            sheet3.write(v, 0, k)
                            sheet3.write(
                                v, 1, round(float(scanpaths[keys[i]][j][7]) * 720)
                            )
                            sheet3.write(
                                v, 2, round(float(scanpaths[keys[i]][j][8]) * 576)
                            )
                            sheet3.write(v, 3, scanpaths[keys[i]][j][5])
                            sheet3.write(v, 4, scanpaths[keys[i]][j][3])

                        else:
                            cc += 1
        if cc >= 4:
            print(f"media{m} {keys[i]} 超出4帧及以上")
        if v > 0 and cc < 4:
            f.save(paramas.fixationimgcoor + "/media" + m + "/" + keys[i] + ".xls")


"""
按顺序执行各个函数
"""

pa = paramas()
process_list = []
# videos=getVideoStart(pa.videostartsplit)
# for i in range(1,34):  #开启5个子进程执行fun1函数
#     p = Process(target=videosplit,args=(pa,i)) #实例化进程对象
#     p.start()
#     process_list.append(p)
#
# for i in process_list:
#     i.join()

# for i in range(1,15):  #开启5个子进程执行fun1函数
#     # if i!=14:
#     p = Process(target=sgy1,args=(pa,0.5,i)) #实例化进程对象
#     p.start()
#     process_list.append(p)
#
# for i in process_list:
#     i.join()

# crudegazepointsplit(pa,14)

# missingsamples(pa)

# for i in range(1,34):
#     videosplit(pa,i)

# if not os.path.exists(f'{pa.gazepointfps}'):
#     os.makedirs(f'{pa.gazepointfps}')
# files=os.listdir(f'{pa.gazepointfps}')

for i in range(
    1, 15
):  # 开启5个子进程执行fun1函数paramas.gazepointfps+'/media'+m+'/'+keys[i]+'.xlsx'
    if not os.path.exists(f"{pa.gazepointfps}/media{i}"):
        os.makedirs(f"{pa.gazepointfps}/media{i}")
    scanpaths = getScanpath(f"{pa.delecrudegazepointimgcoor}0.5/media{i}")
    videotime = getVideoStart(pa.videoendsplit)
    p = Process(
        target=gazepointfps, args=(pa, scanpaths, videotime, str(i))
    )  # 实例化进程对象
    # filee='project'+str(i)+'.xlsx'
    # p = Process(target=gazepointcoor,args=(pa,str(i))) #实例化进程对象
    p.start()
    process_list.append(p)

for i in process_list:
    i.join()


"""
def calsaccadecoor(paramas,file):
    sheet=pd.read_excel(f'{paramas.savefile}/{file}',header=None)
    sheet=sheet.values.tolist()
    add=[]
    for i in range(len(sheet[0])):
        add.append('nan')
    sheet.append(add)

    coorx,coory,co=[],[],[]
    nan=-1  #非Saccade
    index=-1
    if sheet[1][57]=='Saccade':
        index=sheet[1][59]
        co.append(1)
        if str(sheet[1][50])!='nan':
            nan=0
            coorx.append(sheet[1][50])
            coory.append(sheet[1][51])
        else:
            nan=1

    for i in range(2,len(sheet)):
        print(i+1)
        if nan==-1:  #上一次记录非Saccade
            coorx, coory, co = [], [], []
            if sheet[i][57]=='Saccade':
                index=sheet[i][59]
                co.append(i)
                if str(sheet[i][50])!='nan':
                    nan=0
                    coorx.append(float(sheet[i][50]))
                    coory.append(float(sheet[i][51]))
                else:
                    nan=1
            else:
                coorx, coory, co = [], [], []
                nan=-1
                index=-1

        elif nan==0:  #上一次记录是Saccade，且坐标不为nan
            if sheet[i][57]=='Saccade' and sheet[i][59]==index:  #与上一次记录属于同一条记录
                co.append(i)
                if str(sheet[i][50])=='nan':
                    print('同一条记录上一个坐标不是nan这一个是')
                    nan=1
                else:
                    coorx.append(float(sheet[i][50]))
                    coory.append(float(sheet[i][51]))
            elif sheet[i][57]=='Saccade' and sheet[i][59]!=index:
                if len(coorx)!=0:
                    newx=np.mean(coorx)
                    newy = np.mean(coory)
                    for j in co:
                        sheet[j][62]=newx
                        sheet[j][63]=newy
                index=sheet[i][59]
                coorx, coory, co = [], [], []
                co.append(i)
                if str(sheet[i][50])=='nan':
                    nan=1
                else:
                    nan=0
                    coorx.append(float(sheet[i][50]))
                    coory.append(float(sheet[i][51]))
            elif sheet[i][57]!='Saccade':
                nan=-1
                index=-1
                if len(coorx)!=0:
                    newx = np.mean(coorx)
                    newy = np.mean(coory)
                    for j in co:
                        sheet[j][62] = newx
                        sheet[j][63] = newy
                coorx, coory, co = [], [], []

        elif nan==1:
            if sheet[i][57]=='Saccade' and sheet[i][59]==index:
                co.append(i)
                if str(sheet[i][50])!='nan':
                    print('同一条记录上一个坐标是nan这一个不是')
                    nan=0
                    coorx.append(float(sheet[i][50]))
                    coory.append(float(sheet[i][51]))
            elif sheet[i][57]=='Saccade' and sheet[i][59]!=index:
                if len(coorx)!=0:
                    newx=np.mean(coorx)
                    newy = np.mean(coory)
                    for j in co:
                        sheet[j][62]=newx
                        sheet[j][63]=newy
                index=sheet[i][59]
                coorx, coory, co = [], [], []
                co.append(i)
                if str(sheet[i][50])=='nan':
                    nan=1
                else:
                    nan=0
                    coorx.append(float(sheet[i][50]))
                    coory.append(float(sheet[i][51]))
            elif sheet[i][57]!='Saccade':
                if len(coorx)!=0:
                    newx=np.mean(coorx)
                    newy = np.mean(coory)
                    for j in co:
                        sheet[j][62]=newx
                        sheet[j][63]=newy
                nan=-1
                coorx, coory, co = [], [], []
                index=-1
    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(sheet) + 1):
        for j in range(1, len(sheet[i - 1]) + 1):
            outws.cell(i, j).value = sheet[i - 1][j - 1]  # 写文件
    outwb.save(f'{paramas.savesaccadecoor}/{file}')  # 一定要记得保存

def calgazepointcoor(paramas,file):
    sheet=pd.read_excel(f'{paramas.savefile}/{file}',header=None)
    sheet=sheet.values.tolist()
    add=[]
    for i in range(len(sheet[0])):
        add.append('nan')
    sheet.append(add)

    coorx,coory,co=[],[],[]
    index=0
    nan=1
    if str(sheet[1][50])!='nan':
        nan=0
        index+=1
        co.append(1)
        coorx.append(sheet[1][50])
        coory.append(sheet[1][51])

    for i in range(2,len(sheet)):
        print(i+1)
        if nan==0:
            if str(sheet[i][50])!='nan':
                index+=1
                co.append(i)
                coorx.append(float(sheet[i][50]))
                coory.append(float(sheet[i][51]))
                if index==5:
                    x=np.mean(coorx)
                    y=np.mean(coory)
                    for j in co:
                        sheet[j][50]=x
                        sheet[j][51]=y
                    index=0
                    coorx,coory,co=[],[],[]
            else:
                nan=1
                if index!=0:
                    x = np.mean(coorx)
                    y = np.mean(coory)
                    for j in co:
                        sheet[j][50] = x
                        sheet[j][51] = y
                index=0
                coorx,coory,co=[],[],[]
        else:
            if str(sheet[i][50])!='nan':
                nan=0
                index+=1
                co.append(i)
                coorx.append(float(sheet[i][50]))
                coory.append(float(sheet[i][51]))

    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(sheet) + 1):
        for j in range(1, len(sheet[i - 1]) + 1):
            outws.cell(i, j).value = sheet[i - 1][j - 1]  # 写文件
    outwb.save(f'{paramas.savegazepointcoor}/{file}')  # 一定要记得保存

def savefixation_saccade(paramas,file):
    sheet = pd.read_excel(f'{paramas.savesaccadecoor}/project{file}.xlsx',header=None)
    sheet = sheet.values.tolist()

    #标记样本是第几次实验
    add=[]
    for i in range(len(sheet[0])):
        add.append('nan')
    sheet.append(add)
    participant=sheet[1][3]
    rou=1
    lenmedia=0

    for i in range(2,len(sheet)):
        if sheet[i][3]==participant:
            if sheet[i][23]=='VideoStimulusStart':
                lenmedia+=1
                if lenmedia>14:
                    rou+=1
                    lenmedia=1
            sheet[i][3]=participant+str(rou)
        else:
            rou=1
            lenmedia=0
            participant=sheet[i][3]

    ##删除无用列
    ##de为要删除的列索引
    de = []
    de.append(2)
    for i in range(4, 22):
        de.append(i)
    for i in range(23, 50):
        de.append(i)
    for i in range(52, 56):
        de.append(i)
    de.extend([60,61])
    sheet = np.delete(sheet, de, 1)
    print(f'file{file} delete column')
    print(len(sheet[0]))

    ##删除无用行
    de=[]
    for i in range(1,len(sheet)):
        if not ('媒体' in sheet[i][6]):
            de.append(i)
    sheet=np.delete(sheet,de,0)
    print(f'file{file} delete row (6)')

    de = []
    for i in range(1, len(sheet)):
        if not (sheet[i][7] in ['Saccade' ,'Fixation']):
            de.append(i)
    sheet=np.delete(sheet,de,0)
    print(f'file{file} delete row (7)')

    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(sheet) + 1):
        for j in range(1, len(sheet[i - 1]) + 1):
            outws.cell(i, j).value = sheet[i - 1][j - 1]  # 写文件
    outwb.save(f'{paramas.savesaccadefixation}/project{file}.xlsx')  # 一定要记得保存

def saveallgazepoint(paramas,file):
    sheet = pd.read_excel(f'{paramas.savegazepointcoor}/project{file}.xlsx',header=None)
    sheet = sheet.values.tolist()

    #标记样本是第几次实验
    add=[]
    for i in range(len(sheet[0])):
        add.append('nan')
    sheet.append(add)
    participant=sheet[1][3]
    rou=1
    lenmedia=0

    for i in range(2,len(sheet)):
        if sheet[i][3]==participant:
            if sheet[i][23]=='VideoStimulusStart':
                lenmedia+=1
                if lenmedia>14:
                    rou+=1
                    lenmedia=1
            sheet[i][3]=participant+str(rou)
        else:
            rou=1
            lenmedia=0
            participant=sheet[i][3]

    ##删除无用列
    ##de为要删除的列索引
    de = []
    de.append(2)
    for i in range(4, 22):
        de.append(i)
    for i in range(23, 50):
        de.append(i)
    for i in range(52, 56):
        de.append(i)
    de.extend([60,61])
    sheet = np.delete(sheet, de, 1)
    print(f'file{file} delete column')
    print(len(sheet[0]))

    ##删除无用行
    de=[]
    for i in range(1,len(sheet)):
        if not ('媒体' in sheet[i][6]):
            de.append(i)
        else:
            if str(sheet[i][4])=='nan':
                de.append(i)
    sheet=np.delete(sheet,de,0)
    print(f'file{file} delete row (4,6)')

    ##保存文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for i in range(1, len(sheet) + 1):
        for j in range(1, len(sheet[i - 1]) + 1):
            outws.cell(i, j).value = sheet[i - 1][j - 1]  # 写文件
    outwb.save(f'{paramas.savegazepoint}/project{file}.xlsx')  # 一定要记得保存

#所有Gaze point
def gazepointderepeat(paramas,file):
    sheet = pd.read_excel(f'{paramas.savegazepoint}/project{file}.xlsx', header=None)
    sheet = sheet.values.tolist()
    add=[]
    for i in range(len(sheet[0])):
        add.append(-1)
    sheet.append(add)
    newsheet=[]
    newsheet.append(['Recording timestamp','Participant name','Gaze point X (MCSnorm)','Gaze point Y (MCSnorm)','Presented Media name',
                     'Eye movement type','Gaze event duration','Eye movement type index','Fixation point X (MCSnorm)',
                     'Fixation point Y (MCSnorm)'])
    newsheet.append([sheet[1][0],sheet[1][2],sheet[1][4],sheet[1][5],sheet[1][6],sheet[1][7],sheet[1][8],sheet[1][9],sheet[1][10]
                     ,sheet[1][11]])

    participant=sheet[1][2]

    for i in range(2,len(sheet)):
        if sheet[i][2]==participant:
            newsheet.append(
                [sheet[i][0], sheet[i][2], sheet[i][4], sheet[i][5], sheet[i][6], sheet[i][7], sheet[i][8], sheet[i][9],
                 sheet[i][10], sheet[i][11]])
        else:
            outwb = openpyxl.Workbook()  # 打开一个将写的文件
            outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
            for k in range(1, len(newsheet) + 1):
                for j in range(1, len(newsheet[k - 1]) + 1):
                    outws.cell(k, j).value = newsheet[k - 1][j - 1]  # 写文件
            outwb.save(f'{paramas.okgazepoint}/{participant}.xlsx')  # 一定要记得保存

            participant=sheet[i][2]
            newsheet=[]
            newsheet.append(
                ['Recording timestamp', 'Participant name', 'Gaze point X (MCSnorm)', 'Gaze point Y (MCSnorm)',
                 'Presented Media name',
                 'Eye movement type', 'Gaze event duration', 'Eye movement type index', 'Fixation point X (MCSnorm)',
                 'Fixation point Y (MCSnorm)'])
            newsheet.append(
                [sheet[i][0], sheet[i][2], sheet[i][4], sheet[i][5], sheet[i][6], sheet[i][7], sheet[i][8], sheet[i][9],
                 sheet[i][10], sheet[i][11]])

##查看缺失样本分布
def missingsamples(paramas):
    if not os.path.exists(paramas.missingsamples):
        os.makedirs(paramas.missingsamples)
    alldele1,alldele2,alldele3,alldele4,alldele5=0,0,0,0,0
    alldeles1, alldeles2, alldeles3, alldeles4, alldeles5 = [],[],[],[],[]
    talldele1, talldele2, talldele3, talldele4, talldele5 = 0, 0, 0, 0, 0
    aalldele1,aalldele2,aalldele3,aalldele4,aalldele5=0,0,0,0,0
    allfile=0
    alldele=0
    for media in range(1,15):
        files=os.listdir(f'{paramas.gazepointimgcoor}/media{media}')
        allfile+=len(files)
        # print(f'media{media}共有{len(files)}个样本')
        dele=0
        dele1,dele2,dele3,dele4,dele5=0,0,0,0,0
        tdele1, tdele2, tdele3, tdele4, tdele5 = 0, 0, 0, 0, 0
        adele1, adele2, adele3, adele4, adele5 = 0, 0, 0, 0, 0
        for file in files:
            sheet = pd.read_excel(f'{paramas.crudegazepointimgcoor}/media{media}/{file[:-4]}.xlsx',header=None)
            sheet=sheet.values.tolist()
            nan=0
            for i in range(1,len(sheet)):
                if str(sheet[i][2])=='nan':
                    nan+=1
            if nan / (len(sheet) - 1) >= 0.5:
                alldele+=1
                dele+=1
            if nan/(len(sheet)-1)>=0.5 and nan/(len(sheet)-1)<0.6:
                dele1+=1
                alldele1+=1
                alldeles1.append(['media'+str(media),file[:-4]])
                if file[0]=='n':
                    tdele1 += 1
                    talldele1 += 1
                else:
                    adele1 += 1
                    aalldele1 += 1
            elif nan/(len(sheet)-1)>=0.6 and nan/(len(sheet)-1)<0.7:
                dele2+=1
                alldele2+=1
                alldeles2.append(['media' + str(media), file[:-4]])
                if file[0]=='n':
                    tdele2 += 1
                    talldele2 += 1
                else:
                    adele2 += 1
                    aalldele2 += 1
            elif nan/(len(sheet)-1)>=0.7 and nan/(len(sheet)-1)<0.8:
                dele3+=1
                alldele3+=1
                alldeles3.append(['media' + str(media), file[:-4]])
                if file[0]=='n':
                    tdele3 += 1
                    talldele3 += 1
                else:
                    adele3 += 1
                    aalldele3 += 1
            elif nan/(len(sheet)-1)>=0.8 and nan/(len(sheet)-1)<0.9:
                dele4+=1
                alldele4+=1
                alldeles4.append(['media' + str(media), file[:-4]])
                if file[0]=='n':
                    tdele4 += 1
                    talldele4 += 1
                else:
                    adele4 += 1
                    aalldele4 += 1
            elif nan/(len(sheet)-1)>=0.9 :
                dele5+=1
                alldele5+=1
                alldeles5.append(['media' + str(media), file[:-4]])
                if file[0]=='n':
                    tdele5 += 1
                    talldele5 += 1
                else:
                    adele5 += 1
                    aalldele5 += 1
        print(f'media{media}缺失记录占比位于[0.5,0.6)的样本有{dele1}/{dele} （{np.round(dele1/dele,4)}），其中ASD占{np.round(adele1/dele1,4)}，TD占{np.round(tdele1/dele1,4)}')
        print(
            f'media{media}缺失记录占比位于[0.6,0.7)的样本有{dele2}/{dele} （{np.round(dele2 / dele, 4)}），其中ASD占{np.round(adele2 / dele2, 4)}，TD占{np.round(tdele2 / dele2, 4)}')
        print(
            f'media{media}缺失记录占比位于[0.7,0.8)的样本有{dele3}/{dele} （{np.round(dele3 / dele, 4)}），其中ASD占{np.round(adele3 / dele3, 4)}，TD占{np.round(tdele3 / dele3, 4)}')
        print(
            f'media{media}缺失记录占比位于[0.8,0.9)的样本有{dele4}/{dele} （{np.round(dele4 / dele, 4)}），其中ASD占{np.round(adele4 / dele4, 4)}，TD占{np.round(tdele4 / dele4, 4)}')
        print(
            f'media{media}缺失记录占比位于[0.9,1)的样本有{dele5}/{dele} （{np.round(dele5 / dele, 4)}），其中ASD占{np.round(adele5 / dele5, 4)}，TD占{np.round(tdele5 / dele5, 4)}')
        print()
    print(
        f'缺失记录占比位于[0.5,0.6)的样本有{alldele1}/{alldele} （{np.round(alldele1 / alldele, 4)}），其中ASD占{np.round(aalldele1 / alldele1, 4)}，TD占{np.round(talldele1 / alldele1, 4)}')
    print(
        f'缺失记录占比位于[0.6,0.7)的样本有{alldele2}/{alldele} （{np.round(alldele2 / alldele, 4)}），其中ASD占{np.round(aalldele2 / alldele2, 4)}，TD占{np.round(talldele2 / alldele2, 4)}')
    print(
        f'缺失记录占比位于[0.7,0.8)的样本有{alldele3}/{alldele} （{np.round(alldele3 / alldele, 4)}），其中ASD占{np.round(aalldele3 / alldele3, 4)}，TD占{np.round(talldele3 / alldele3, 4)}')
    print(
        f'缺失记录占比位于[0.8,0.9)的样本有{alldele4}/{alldele} （{np.round(alldele4 / alldele, 4)}），其中ASD占{np.round(aalldele4 / alldele4, 4)}，TD占{np.round(talldele4 / alldele4, 4)}')
    print(
        f'缺失记录占比位于[0.9,1)的样本有{alldele5}/{alldele} （{np.round(alldele5 / alldele, 4)}），其中ASD占{np.round(aalldele5 / alldele5, 4)}，TD占{np.round(talldele5 / alldele5, 4)}')
    alldeles=[]
    alldeles.append(alldeles1)
    alldeles.append(alldeles2)
    alldeles.append(alldeles3)
    alldeles.append(alldeles4)
    alldeles.append(alldeles5)
    for i in range(5):
        f=open(f'{paramas.missingsamples}/{i+1}.txt','w')
        for j in range(len(alldeles[i])):
            for k in range(len(alldeles[i][j])):
                f.write(alldeles[i][j][k])
                f.write('\t')
            f.write('\n')
        f.close()


##删掉表第一行，只保留时间列和xy的表格，并且时间列的数字变成顺序123这样
def sgy1(paramas,sca,media):
    if not os.path.exists(f'{paramas.xy}/media{media}'):
        os.makedirs(f'{paramas.xy}/media{media}')
    files = os.listdir(f'{paramas.delecrudegazepointimgcoor}{sca}/media{media}')
    for file in files:
        newsheet=[]
        sheet = pd.read_excel(f'{paramas.delecrudegazepointimgcoor}{sca}/media{media}/{file}',header=None)
        sheet = sheet.values.tolist()
        for i in range(1, len(sheet)):
            if str(sheet[i][2])!='nan':
                newsheet.append([i,round(float(sheet[i][2])*720),round(float(sheet[i][3])*576)])
            else:
                newsheet.append([i,sheet[i][2],sheet[i][3]])
        outwb = openpyxl.Workbook()  # 打开一个将写的文件
        outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
        for k in range(1, len(newsheet) + 1):
            for j in range(1, len(newsheet[k - 1]) + 1):
                outws.cell(k, j).value = newsheet[k - 1][j - 1]  # 写文件
        outwb.save(f'{paramas.xy}/media{media}/{file}')  # 一定要记得保存
"""
