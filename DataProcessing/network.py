import numpy as np
import os
import pandas as pd
from matplotlib.path import Path
import math
import cv2
import openpyxl
from multiprocessing import Process
import xlwt
from numba import jit
import networkx as nx

##生成显著性拼接图像

size = {'media1': 234, 'media2': 286, 'media3':44,'media4':225, 'media5': 186, 'media6':64,'media7': 80,
            'media8': 83, 'media9': 152,'media10': 128, 'media11': 135,'media12':58, 'media13':75,'media14':86}

class paramas():
    def __init__(self):
        self.datafile = '20210623'
        self.savefile = '1datadecolumn'

        self.savesaccadecoor = '2saccadecoor'
        self.savegazepointcoor = '2gazepointcoor'

        self.savesaccadefixation = '3saccadefixation'
        self.savefixation = '3fixation'
        self.savegazepoint = '3gazepoint'
        self.videostart = 'videostart'

        self.oksaccadefixation = '5oksaccadefixation'
        self.okfixation = '5okfixation'
        self.okgazepoint = '5okgazepoint'
        self.videostartsplit = 'videostartsplit'

        self.fixationsample = '6fixationsample'
        self.saccadefixationsample = '6saccadefixationsample'
        self.gazepointsample = '6gazepointsample'

        self.fixationimgcoor = 'fixationimg/fixationcoor'
        self.saccadefixationimgcoor = 'saccadefixationimg/fixationcoor'
        self.gazepointimgcoor = 'gazepointimg/fixationcoor'
        self.gazepointfps = 'gazepointfps'

        self.fixationfsjoint = 'fixationimg/fsjoint14ss'
        self.saccadefixationfsjoint = 'saccadefixationimg/fsjoint10ss'
        self.gazepointfsjoint = 'gazepointimg/fsjoint14ss'
        self.newfixationfsjoint = 'newfixationimg/fsjoint10ss'
        self.newsaccadefixationfsjoint = 'newsaccadefixationimg/fsjoint10ss'
        self.newgazepointfsjoint = 'newgazepointimg/fsjoint10ss'
        self.saliency='../Saliency'
        self.newsaliency='../TASED-Net-master/output'
        self.frame='../img/frame'
        self.gazepointfsjoint1d='gazepointimg/fsjoint1d'
        self.newgazepointfsjoint1d = 'newgazepointimg/fsjoint1d'
        self.fixationfsjoint1d='fixationimg/fsjoint1d'
        self.newfixationfsjoint1d='newfixationimg/fsjoint1d'
        self.saccadefixationfsjoint1d='saccadefixationimg/fsjoint1d'
        self.newsaccadefixationfsjoint1d='newsaccadefixationimg/fsjoint1d'

        self.framecoor='framecoordinate'
        self.visualelement='visualelement'
        self.firstfixationtime='stafeatures/firstfixationtime'
        self.aoifeatures='stafeatures/aoi'
        self.stafeatures='stafeatures/stafeatures'
        self.netfeatures='stafeatures/netfeatures'
        self.stanetfeatures = 'stafeatures/stanetfeatures'
        self.stapath = 'stapath'
        self.similarity='stafeatures/similarity'

def element(paramas,file):
    if not os.path.exists(f'{paramas.visualelement}/media{file}'):
        os.makedirs(f'{paramas.visualelement}/media{file}')
    landmark=pd.read_excel(f'{paramas.framecoor}/媒体{file}.xlsx',header=None)
    df=landmark.values.tolist()
    samples=os.listdir(f'{paramas.fixationimgcoor}/media{file}')
    for sample in samples:
        e=0
        individual_path=[]
        fixations=pd.read_excel(f'{paramas.fixationsample}/{sample[:-4]}.xlsx')
        fixations=fixations.values.tolist()
        start=pd.read_excel(f'{paramas.videostartsplit}/{sample[:-4]}.xlsx',header=None)
        start=start.values.tolist()
        for i in range(len(start)):
            if start[i][2][2:-4]==str(file):
                T0=int(start[i][0])
                break

        for co in fixations:
            if co[2][2:-4]==str(file) and str(co[7])!='nan':
                fps = (int(co[0]) - T0) // 40000
                if fps==0 or fps>size['media'+str(file)]:
                    pass
                else:
                    e+=1
                    eye1x=int(df[fps][10])
                    eye1y=int(df[fps][11])
                    eye2x=int(df[fps][12])
                    eye2y=int(df[fps][13])
                    eye3x=int(df[fps][14])
                    eye3y=int(df[fps][15])
                    eye4x=int(df[fps][16])
                    eye4y=int(df[fps][17])
                    eye=Path([(eye1x,eye1y),(eye2x,eye2y),(eye3x,eye3y),(eye4x,eye4y)])
                    nose1x=df[fps][18]
                    nose1y=df[fps][19]
                    nose2x=df[fps][20]
                    nose2y=df[fps][21]
                    nose3x=df[fps][22]
                    nose3y=df[fps][23]
                    nose4x=df[fps][24]
                    nose4y=df[fps][25]
                    nose=Path([(nose1x,nose1y),(nose2x,nose2y),(nose3x,nose3y),(nose4x,nose4y)])
                    mouth1x=df[fps][26]
                    mouth1y=df[fps][27]
                    mouth2x=df[fps][28]
                    mouth2y=df[fps][29]
                    mouth3x=df[fps][30]
                    mouth3y=df[fps][31]
                    mouth4x=df[fps][32]
                    mouth4y=df[fps][33]
                    mouth=Path([(mouth1x,mouth1y),(mouth2x,mouth2y),(mouth3x,mouth3y),(mouth4x,mouth4y)])
                    centerX=df[fps][1]
                    centerY = df[fps][2]
                    short=math.ceil(df[fps][7])
                    long=math.ceil(df[fps][8])
                    angle=df[fps][9]

                    co[7]=round(float(co[7])*720)
                    co[8] = round(float(co[8]) * 576)
                    x=co[7]
                    y=co[8]
                    coo=np.array((x,y))
                # coo[0][0]=x
                # coo[0][1]=y
                    coo.reshape(1,2)
                    if math.pow(((x-centerX)*math.cos(angle)+(y-centerY)*math.sin(angle))/long,2)+math.pow(((centerX-x)*math.sin(angle)+(y-centerY)*math.cos(angle))/short,2)<=1:
                        face_in=True
                    else:
                        face_in=False
                    eye_in = eye.contains_point(coo)
                    nose_in=nose.contains_point(coo)
                    mouth_in=mouth.contains_point(coo)

                    if not(face_in):
                        result='E'
                    elif face_in and not(eye_in or nose_in or mouth_in):
                        result='D'
                    elif face_in and eye_in:
                        if not(nose_in):
                            result='A'
                        else:
                            result='A'
                    elif face_in and nose_in:
                        if not(mouth_in):
                            result='B'
                        else:
                            result='B'
                    elif face_in and mouth_in:
                        result='C'
                    # else:
                    #     result='error'

                    co.append(result)
                    individual_path.append(co)
        if e>1:
            f = open(f'{paramas.visualelement}/media{file}/{sample[:-4]}.txt', 'w')
            f.write('RecordTime'+'\t'+'Participant'+'\t'+'Media'+'\t'+'EyeMovement'+'\t'+'EventDuration'+'\t'+'Duration'+'\t'+'EventIndex'+'\t'+'coordinateX'+'\t'+'coordinateY'+'\t'+'AOI'+'\n')
            for i in range(len(individual_path)):
                for j in range(len(individual_path[i])):
                    f.write(str(individual_path[i][j]))  # write函数不能写int类型的参数，所以使用str()转化
                    f.write('\t')  # 相当于Tab一下，换一个单元格
                f.write('\n')  # 写完一行立马换行
                # f.writelines(str(individual_path))
            f.close()

def Time_to_first_view(paramas,file):
    if not os.path.exists(f'{paramas.firstfixationtime}'):
        os.makedirs(f'{paramas.firstfixationtime}')
    samples=os.listdir(paramas.fixationsample)
    features=[]
    video=round(size['media'+str(file)]*40)
    mint=np.inf
    maxt=0
    for sample in samples:
        videostart=pd.read_excel(f'{paramas.videostartsplit}/{sample}',header=None)
        videostart=videostart.values.tolist()
        for i in range(len(videostart)):
            if videostart[i][2][2:-4]==str(file):
                start=int(videostart[i][0])
                break
        sheet2=pd.read_excel(f'{paramas.fixationsample}/{sample}')
        sheet2=sheet2.values.tolist()
        for sh in sheet2:
            if sh[2][2:-4]==str(file):
                firstTime=round((int(sh[0])-start)/1000/video,4)
                if firstTime>maxt:
                    maxt=firstTime
                if firstTime<mint:
                    mint=firstTime
                features.append([sample.split('.')[0],firstTime])
                break
    for i in range(len(features)):
        features[i].append(round((features[i][1]-mint)/(maxt-mint),4))
    f = xlwt.Workbook()
    sheet3 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    sheet3.write(0, 0, 'sample')
    sheet3.write(0, 1, 'FirstFixationTime')
    sheet3.write(0, 2, 'FirstFixationTime(normalization)')
    for i in range(len(features)):
        for j in range(len(features[i])):
            sheet3.write(i+1,j,features[i][j])
    f.save(f'{paramas.firstfixationtime}/media{file}.xls')

def stastic_features(paramas,file):
    if not os.path.exists(f'{paramas.aoifeatures}'):
        os.makedirs(f'{paramas.aoifeatures}')
    samples = os.listdir(f'{paramas.visualelement}/media{file}')
    features = []
    minfixations, minduration, minm_duration, minsd_duration, minfixa, minfixb, minfixc, minfixd, minfixe, mindura, mindurb, mindurc, \
    mindurd, mindure, minmdisimg, minsddisimg, minmdisscan, minsddisscan,minrevisits =10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9
    maxfixations, maxduration, maxm_duration, maxsd_duration, maxfixa, maxfixb, maxfixc, maxfixd, maxfixe, maxdura, maxdurb, maxdurc, \
    maxdurd, maxdure, maxmdisimg, maxsddisimg, maxmdisscan, maxsddisscan,maxrevisits =0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
    for sample in samples:
        fo = open(f'{paramas.visualelement}/media{file}/{sample}', "r")
        myFile = fo.read()
        myRecords = myFile.split('\n')
        myRecords_templist = []
        for y in range(1, len(myRecords) - 1):
            myRecords_templist.append(myRecords[y].split('\t'))  # 把表格数据存入
        duration,m_duration,sd_duration,fixa,fixb,fixc,fixd,fixe,dura,durb,durc,durd,dure,mdisimg,sddisimg,mdisscan,\
            sddisscan,revisits= 0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0
        disimg,disscan,fixcoor,record = [],[],[],[]
        k = 0  # 当前为第k条记录
        m = 0  # m=0表示
        fixations=len(myRecords_templist)
        for data in myRecords_templist:
            duration+=int(data[5])
            record.append(int(data[5]))
            fixcoor.append([float(data[7]),float(data[8])])
            if data[9] == 'A':
                fixa += 1
                dura += int(data[5])
            elif data[9] == 'B':
                fixb += 1
                durb += int(data[5])
            elif data[9] == 'C':
                fixc += 1
                durc += int(data[5])
            elif data[9] == 'D':
                fixd += 1
                durd += int(data[5])
            elif data[9] == 'E':
                fixe += 1
                dure += int(data[5])
        for j in range(len(fixcoor)):
            disimg.append(math.sqrt(math.pow((fixcoor[j][0] - 360), 2) + math.pow((fixcoor[j][1] - 288), 2)))
        mdisimg = np.mean(disimg)
        sddisimg = np.var(disimg)
        if fixa>1:
            revisits=revisits+fixa-1
        if fixb>1:
            revisits=revisits+fixb-1
        if fixc>1:
            revisits=revisits+fixc-1
        if fixd>1:
            revisits=revisits+fixd-1
        if fixe>1:
            revisits=revisits+fixe-1
        if len(fixcoor) % 2 == 0:
            index = int(len(fixcoor) / 2) - 1
            x = (fixcoor[index][0] + fixcoor[index + 1][0]) / 2
            y = (fixcoor[index][1] + fixcoor[index + 1][1]) / 2
            for j in range(len(fixcoor)):
                disscan.append(math.sqrt(math.pow((fixcoor[j][0] - x), 2) + math.pow((fixcoor[j][1] - y), 2)))
            mdisscan = np.mean(disscan)
            sddisscan = np.var(disscan)
        else:
            if len(fixcoor) == 1:
                mdisscan = 0
                sddisscan = 0
            else:
                index = len(fixcoor) // 2
                x = fixcoor[index][0]
                y = fixcoor[index][1]
                for j in range(len(fixcoor)):
                    disscan.append(
                        math.sqrt(math.pow((fixcoor[j][0] - x), 2) + math.pow((fixcoor[j][1] - y), 2)))
                mdisscan = np.mean(disscan)
                sddisscan = np.var(disscan)
        m_duration = np.mean(record)
        sd_duration = np.var(record)

        if fixations>maxfixations:
            maxfixations=fixations
        if fixations<minfixations:
            minfixations=fixations
        if duration > maxduration:
            maxduration = duration
        if duration < minduration:
            minduration = duration
        if m_duration>maxm_duration:
            maxm_duration=m_duration
        if m_duration<minm_duration:
            minm_duration=m_duration
        if sd_duration>maxsd_duration:
            maxsd_duration=sd_duration
        if sd_duration<minsd_duration:
            minsd_duration=sd_duration
        if fixa>maxfixa:
            maxfixa=fixa
        if fixa<minfixa:
            minfixa=fixa
        if fixb > maxfixb:
            maxfixb = fixb
        if fixb < minfixb:
            minfixb = fixb
        if fixc > maxfixc:
            maxfixc = fixc
        if fixc < minfixc:
            minfixc= fixc
        if fixd > maxfixd:
            maxfixd = fixd
        if fixd < minfixd:
            minfixd = fixd
        if fixe > maxfixe:
            maxfixe = fixe
        if fixe < minfixe:
            minfixe = fixe
        if dura>maxdura:
            maxdura=dura
        if dura<mindura:
            mindura=dura
        if durb > maxdurb:
            maxdurb = durb
        if durb < mindurb:
            mindurb = durb
        if durc > maxdurc:
            maxdurc = durc
        if durc < mindurc:
            mindurc= durc
        if durd > maxdurd:
            maxdurd = durd
        if durd < mindurd:
            mindurd = durd
        if dure > maxdure:
            maxdure = dure
        if dure < mindure:
            mindure = dure
        if mdisimg>maxmdisimg:
            maxmdisimg=mdisimg
        if mdisimg<minmdisimg:
            minmdisimg=mdisimg
        if sddisimg>maxsddisimg:
            maxsddisimg=sddisimg
        if sddisimg<minsddisimg:
            minsddisimg=sddisimg
        if mdisscan>maxmdisscan:
            maxmdisscan=mdisscan
        if mdisscan<minmdisscan:
            minmdisscan=mdisscan
        if sddisscan>maxsddisscan:
            maxsddisscan=sddisscan
        if sddisscan<minsddisscan:
            minsddisscan=sddisscan
        if revisits>maxrevisits:
            maxrevisits=revisits
        if revisits<minrevisits:
            minrevisits=revisits

        features.append([sample.split('.')[0],fixations,duration,m_duration,sd_duration,fixa,fixb,fixc,fixd,fixe,dura,durb,durc,durd,
                         dure,mdisimg,sddisimg,mdisscan,sddisscan,revisits])
    f = xlwt.Workbook()
    sheet3 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    line1=['sample','fixations','duration','m_duration','sd_duration','fixa','fixb','fixc','fixd','fixe','dura','durb','durc','durd',
           'dure','mdisimg','sddisimg','mdisscan','sddisscan','revisits','gyhfixations','gyhduration','gyhm_duration','gyhsd_duration','gyhfixa',
           'gyhfixb','gyhfixc','gyhfixd','gyhfixe','gyhdura','gyhdurb','gyhdurc','gyhdurd','gyhdure',
           'gyhmdisimg','gyhsddisimg','gyhmdisscan','gyhsddisscan','gyhrevisits']
    for i in range(len(line1)):
        sheet3.write(0,i,line1[i])

    for i in range(len(features)):
        gyhfixations=np.round((features[i][1]-minfixations)/(maxfixations-minfixations),4)
        gyhduration = np.round((features[i][2] - minduration) / (maxduration - minduration), 4)
        gyhm_duration = np.round((features[i][3] - minm_duration) / (maxm_duration - minm_duration), 4)
        gyhsd_duration = np.round((features[i][4] - minsd_duration) / (maxsd_duration - minsd_duration), 4)
        gyhfixa = np.round((features[i][5] - minfixa) / (maxfixa - minfixa), 4)
        gyhfixb = np.round((features[i][6] - minfixb) / (maxfixb - minfixb), 4)
        gyhfixc = np.round((features[i][7] - minfixc) / (maxfixc - minfixc), 4)
        gyhfixd = np.round((features[i][8] - minfixd) / (maxfixd - minfixd), 4)
        gyhfixe = np.round((features[i][9] - minfixe) / (maxfixe - minfixe), 4)
        gyhdura = np.round((features[i][10] - mindura) / (maxdura - mindura), 4)
        gyhdurb = np.round((features[i][11] - mindurb) / (maxdurb - mindurb), 4)
        gyhdurc = np.round((features[i][12] - mindurc) / (maxdurc - mindurc), 4)
        gyhdurd = np.round((features[i][13] - mindurd) / (maxdurd - mindurd), 4)
        gyhdure = np.round((features[i][14] - mindure) / (maxdure - mindure), 4)
        gyhmdisimg = np.round((features[i][15] - minmdisimg) / (maxmdisimg - minmdisimg), 4)
        gyhsddisimg = np.round((features[i][16] - minsddisimg) / (maxsddisimg - minsddisimg), 4)
        gyhmdisscan = np.round((features[i][17] - minmdisscan) / (maxmdisscan - minmdisscan), 4)
        gyhsddisscan = np.round((features[i][18] - minsddisscan) / (maxsddisscan - minsddisscan), 4)
        gyhrevisits = np.round((features[i][19] - minrevisits) / (maxrevisits - minrevisits), 4)
        features[i].extend([gyhfixations,gyhduration,gyhm_duration,gyhsd_duration,gyhfixa,gyhfixb,gyhfixc,gyhfixd,
                            gyhfixe,gyhdura,gyhdurb,gyhdurc,gyhdurd,gyhdure,gyhmdisimg,gyhsddisimg,gyhmdisscan,gyhsddisscan,gyhrevisits])
        for j in range(len(features[i])):
            sheet3.write(i+1,j,features[i][j])
    f.save(f'{paramas.aoifeatures}/media{file}.xls')

def getStringEditDistance(Sequence1, Sequence2):
    distance = 0
    matrix = []

    for k in range(0, len(Sequence1) + 1):
        matrix.append([])
        for g in range(0, len(Sequence2) + 1):
            matrix[k].append(0)

    for k in range(0, len(Sequence1) + 1):
        matrix[k][0] = k

    for g in range(0, len(Sequence2) + 1):
        matrix[0][g] = g

    for g in range(1, len(Sequence2) + 1):
        for k in range(1, len(Sequence1) + 1):
            if Sequence1[k - 1] == Sequence2[g - 1]:
                matrix[k][g] = min(matrix[k - 1][g - 1] + 0, matrix[k][g - 1] + 1, matrix[k - 1][g] + 1)
            else:
                matrix[k][g] = min(matrix[k - 1][g - 1] + 1, matrix[k][g - 1] + 1, matrix[k - 1][g] + 1)
    distance = matrix[len(Sequence1)][len(Sequence2)]
    return distance

def similarity(paramas,file):
    if not os.path.exists(paramas.similarity):
        os.makedirs(paramas.similarity)
    samples=os.listdir(f'{paramas.visualelement}/media{file}')
    trend=pd.read_excel(f'{paramas.stapath}/trendPath.xls',header=None)
    trend=trend.values.tolist()
    trendtime=pd.read_excel(f'{paramas.stapath}/trendPathTime.xls',header=None)
    trendtime=trendtime.values.tolist()
    tdtrend=trend[file][1]
    Dtd=len(tdtrend)
    asdtrend=trend[file][3]
    Dasd=len(asdtrend)
    tdtrendtime=trendtime[file][1]
    Dtdt=len(tdtrendtime)
    asdtrendtime=trendtime[file][3]
    Dasdt=len(asdtrendtime)
    features=[]
    maxstd,maxsasd,maxstdt,maxsasdt=0,0,0,0
    minstd,minsasd,minstdt,minsasdt=np.inf,np.inf,np.inf,np.inf
    f = xlwt.Workbook()
    sheet3 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    line1 = ['sample', 'std','sasd','stdt','sasdt','gyhstd','gyhsasd','gyhstdt','gyhsasdt']
    for i in range(len(line1)):
        sheet3.write(0, i, line1[i])
    for sample in samples:
        path=''
        pathtime=''
        fo = open(f'{paramas.visualelement}/media{file}/{sample}', "r")
        myFile = fo.read()
        myRecords = myFile.split('\n')
        myRecords_templist = []
        for y in range(1, len(myRecords) - 1):
            myRecords_templist.append(myRecords[y].split('\t'))  # 把表格数据存入
        for fix in myRecords_templist:
            path=path+fix[9]
            repeat=round(int(fix[5])/50)
            for i in range(repeat):
                pathtime=pathtime+fix[9]
        l=len(path)
        lt=len(pathtime)
        td=getStringEditDistance(tdtrend,path)
        if Dtd>l:
            ltd=Dtd
        else:
            ltd=l
        std=1-td/ltd
        asd=getStringEditDistance(asdtrend,path)
        if Dasd>l:
            lasd=Dasd
        else:
            lasd=l
        sasd=1-asd/lasd
        tdt=getStringEditDistance(tdtrendtime,pathtime)
        if Dtdt>lt:
            ltdt=Dtdt
        else:
            ltdt=lt
        stdt=1-tdt/ltdt
        asdt=getStringEditDistance(asdtrendtime,pathtime)
        if Dasdt>lt:
            lasdt=Dasdt
        else:
            lasdt=lt
        sasdt=1-asdt/lasdt
        features.append([sample.split('.')[0],std,sasd,stdt,sasdt])
        if std>maxstd:
            maxstd=std
        if std<minstd:
            minstd=std
        if sasd>maxsasd:
            maxsasd=sasd
        if sasd<minsasd:
            minsasd=sasd
        if stdt>maxstdt:
            maxstdt=stdt
        if stdt<minstdt:
            minstdt=stdt
        if sasdt>maxsasdt:
            maxsasdt=sasdt
        if sasdt<minsasdt:
            minsasdt=sasdt
    for i in range(len(features)):
        gyhstd=np.round((features[i][1]-minstd)/(maxstd-minstd),4)
        gyhsasd = np.round((features[i][2] - minsasd) / (maxsasd - minsasd), 4)
        gyhstdt = np.round((features[i][3] - minstdt) / (maxstdt - minstdt), 4)
        gyhsasdt = np.round((features[i][4] - minsasdt) / (maxsasdt - minsasdt), 4)
        features[i].extend([gyhstd,gyhsasd,gyhstdt,gyhsasdt])
        for j in range(len(features[i])):
            sheet3.write(i + 1, j, features[i][j])
    f.save(f'{paramas.similarity}/media{file}.xls')

def net_features(paramas,file):
    if not os.path.exists(f'{paramas.netfeatures}'):
        os.makedirs(f'{paramas.netfeatures}')
    samples = os.listdir(f'{paramas.visualelement}/media{file}')
    features = []
    mincA, mincB, mincC, mincD, mincE, mindA, mindB, mindC, mindD, mindE, minbA, minbB, minbC, minbD, minbE, mineA, mineB, mineC,\
    mineD,mineE =10e9,10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9, 10e9
    maxcA, maxcB, maxcC, maxcD, maxcE, maxdA, maxdB, maxdC, maxdD, maxdE, maxbA, maxbB, maxbC, maxbD, maxbE, maxeA, maxeB, maxeC, \
    maxeD, maxeE =0,0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
    for sample in samples:
        fo = open(f'{paramas.visualelement}/media{file}/{sample}', "r")
        myFile = fo.read()
        myRecords = myFile.split('\n')
        myRecords_templist = []
        for y in range(1, len(myRecords) - 1):
            myRecords_templist.append(myRecords[y].split('\t'))  # 把表格数据存入
        G = nx.DiGraph()
        G.add_nodes_from(['A', 'B', 'C', 'D', 'E'])
        lastnode=myRecords_templist[0][9]
        for i in range(1,len(myRecords_templist)):
            if myRecords_templist[i][9]!=lastnode:
                G.add_edge(lastnode,myRecords_templist[i][9])
                lastnode=myRecords_templist[i][9]
        c_degree = nx.degree_centrality(G)
        c_degree = list(c_degree.values())

        c_eigenvector = nx.eigenvector_centrality_numpy(G)
        c_eigenvector = list(c_eigenvector.values())

        c_closeness = nx.closeness_centrality(G)
        c_closeness = list(c_closeness.values())

        c_betweenness = nx.betweenness_centrality(G)
        c_betweenness = list(c_betweenness.values())

        if c_betweenness[0]>maxbA:
            maxbA=c_betweenness[0]
        if c_betweenness[1]>maxbB:
            maxbB=c_betweenness[1]
        if c_betweenness[2]>maxbC:
            maxbC=c_betweenness[2]
        if c_betweenness[3]>maxbD:
            maxbD=c_betweenness[3]
        if c_betweenness[4]>maxbE:
            maxbE=c_betweenness[4]
        if c_betweenness[0]<minbA:
            minbA=c_betweenness[0]
        if c_betweenness[1]<minbB:
            minbB=c_betweenness[1]
        if c_betweenness[2]<minbC:
            minbC=c_betweenness[2]
        if c_betweenness[3]<minbD:
            minbD=c_betweenness[3]
        if c_betweenness[4]<minbE:
            minbE=c_betweenness[4]

        if c_degree[0]>maxdA:
            maxdA=c_degree[0]
        if c_degree[1]>maxdB:
            maxdB=c_degree[1]
        if c_degree[2]>maxdC:
            maxdC=c_degree[2]
        if c_degree[3]>maxdD:
            maxdD=c_degree[3]
        if c_degree[4]>maxdE:
            maxdE=c_degree[4]
        if c_degree[0]<mindA:
            mindA=c_degree[0]
        if c_degree[1]<mindB:
            mindB=c_degree[1]
        if c_degree[2]<mindC:
            mindC=c_degree[2]
        if c_degree[3]<mindD:
            mindD=c_degree[3]
        if c_degree[4]<mindE:
            mindE=c_degree[4]

        if c_eigenvector[0]>maxeA:
            maxeA=c_eigenvector[0]
        if c_eigenvector[1]>maxeB:
            maxeB=c_eigenvector[1]
        if c_eigenvector[2]>maxeC:
            maxeC=c_eigenvector[2]
        if c_eigenvector[3]>maxeD:
            maxeD=c_eigenvector[3]
        if c_eigenvector[4]>maxeE:
            maxeE=c_eigenvector[4]
        if c_eigenvector[0]<mineA:
            mineA=c_eigenvector[0]
        if c_eigenvector[1]<mineB:
            mineB=c_eigenvector[1]
        if c_eigenvector[2]<mineC:
            mineC=c_eigenvector[2]
        if c_eigenvector[3]<mineD:
            mineD=c_eigenvector[3]
        if c_eigenvector[4]<mineE:
            mineE=c_eigenvector[4]

        if c_closeness[0]>maxcA:
            maxcA=c_closeness[0]
        if c_closeness[1]>maxcB:
            maxcB=c_closeness[1]
        if c_closeness[2]>maxcC:
            maxcC=c_closeness[2]
        if c_closeness[3]>maxcD:
            maxcD=c_closeness[3]
        if c_closeness[4]>maxcE:
            maxcE=c_closeness[4]
        if c_closeness[0]<mincA:
            mincA=c_closeness[0]
        if c_closeness[1]<mincB:
            mincB=c_closeness[1]
        if c_closeness[2]<mincC:
            mincC=c_closeness[2]
        if c_closeness[3]<mincD:
            mincD=c_closeness[3]
        if c_closeness[4]<mincE:
            mincE=c_closeness[4]

        sfea=[sample.split('.')[0]]
        sfea.extend(c_degree)
        sfea.extend(c_betweenness)
        sfea.extend(c_eigenvector)
        sfea.extend(c_closeness)

        features.append(sfea)
    f = xlwt.Workbook()
    sheet3 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    line1=['sample','degree_centrality_A','degree_centrality_B','degree_centrality_C','degree_centrality_D','degree_centrality_E','betweenness_centrality_A'
           ,'betweenness_centrality_B','betweenness_centrality_C','betweenness_centrality_D','betweenness_centrality_E','eigenvector_centrality_A'
           ,'eigenvector_centrality_B','eigenvector_centrality_C','eigenvector_centrality_D','eigenvector_centrality_E','closeness_centrality_A'
           ,'closeness_centrality_B','closeness_centrality_C','closeness_centrality_D','closeness_centrality_E','gyh_degree_centrality_A',
           'gyh_degree_centrality_B','gyh_degree_centrality_C','gyh_degree_centrality_D','gyh_degree_centrality_E','gyh_betweenness_centrality_A'
           ,'gyh_betweenness_centrality_B','gyh_betweenness_centrality_C','gyh_betweenness_centrality_D','gyh_betweenness_centrality_E','gyh_eigenvector_centrality_A'
           ,'gyh_eigenvector_centrality_B','gyh_eigenvector_centrality_C','gyh_eigenvector_centrality_D','gyh_eigenvector_centrality_E','gyh_closeness_centrality_A'
           ,'gyh_closeness_centrality_B','gyh_closeness_centrality_C','gyh_closeness_centrality_D','gyh_closeness_centrality_E']
    for i in range(len(line1)):
        sheet3.write(0,i,line1[i])

    for i in range(len(features)):
        gyhda=np.round((features[i][1]-mindA)/(maxdA-mindA),4)
        gyhdb = np.round((features[i][2] - mindB) / (maxdB - mindB), 4)
        gyhdc = np.round((features[i][3] - mindC) / (maxdC - mindC), 4)
        gyhdd = np.round((features[i][4] - mindD) / (maxdD - mindD), 4)
        gyhde = np.round((features[i][5] - mindE) / (maxdE - mindE), 4)
        gyhba = np.round((features[i][6] - minbA) / (maxbA - minbA), 4)
        gyhbb = np.round((features[i][7] - minbB) / (maxbB - minbB), 4)
        gyhbc = np.round((features[i][8] - minbC) / (maxbC - minbC), 4)
        gyhbd = np.round((features[i][9] - minbD) / (maxbD - minbD), 4)
        gyhbe = np.round((features[i][10] - minbE) / (maxbE - minbE), 4)
        gyhea = np.round((features[i][11] - mineA) / (maxeA - mineA), 4)
        gyheb = np.round((features[i][12] - mineB) / (maxeB - mineB), 4)
        gyhec = np.round((features[i][13] - mineC) / (maxeC - mineC), 4)
        gyhed = np.round((features[i][14] - mineD) / (maxeD - mineD), 4)
        gyhee = np.round((features[i][15] - mineE) / (maxeE - mineE), 4)
        gyhca = np.round((features[i][16] - mincA) / (maxcA - mincA), 4)
        gyhcb = np.round((features[i][17] - mincB) / (maxcB - mincB), 4)
        gyhcc = np.round((features[i][18] - mincC) / (maxcC - mincC), 4)
        gyhcd = np.round((features[i][19] - mincD) / (maxcD - mincD), 4)
        gyhce = np.round((features[i][20] - mincE) / (maxcE - mincE), 4)
        features[i].extend([gyhda,gyhdb,gyhdc,gyhdd,gyhde,gyhba,gyhbb,gyhbc,gyhbd,gyhbe,gyhea,gyheb,gyhec,gyhed,gyhee,gyhca,gyhcb,gyhcc,gyhcd,gyhce])
        for j in range(len(features[i])):
            sheet3.write(i+1,j,features[i][j])
    f.save(f'{paramas.netfeatures}/media{file}.xls')

def joint(paramas,file):
    if not os.path.exists(paramas.stanetfeatures):
        os.makedirs(paramas.stanetfeatures)
    features=pd.read_excel(f'{paramas.aoifeatures}/media{file}.xls')
    features=features.values.tolist()
    firsttime=pd.read_excel(f'{paramas.firstfixationtime}/media{file}.xls')
    firsttime=firsttime.values.tolist()
    similarity=pd.read_excel(f'{paramas.similarity}/media{file}.xls')
    similarity=similarity.values.tolist()
    network = pd.read_excel(f'{paramas.netfeatures}/media{file}.xls')
    network = network.values.tolist()
    f = xlwt.Workbook()
    sheet3 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    line1 = ['sample', 'fixations', 'duration', 'm_duration', 'sd_duration', 'fixa', 'fixb', 'fixc', 'fixd', 'fixe','dura', 'durb',
             'durc', 'durd','dure', 'mdisimg', 'sddisimg', 'mdisscan', 'sddisscan', 'revisits','firstfixationtime','std','sasd',
             'stdt','sasdt','degree_centrality_A','degree_centrality_B','degree_centrality_C','degree_centrality_D','degree_centrality_E','betweenness_centrality_A'
           ,'betweenness_centrality_B','betweenness_centrality_C','betweenness_centrality_D','betweenness_centrality_E','eigenvector_centrality_A'
           ,'eigenvector_centrality_B','eigenvector_centrality_C','eigenvector_centrality_D','eigenvector_centrality_E','closeness_centrality_A'
           ,'closeness_centrality_B','closeness_centrality_C','closeness_centrality_D','closeness_centrality_E',
             'gyhfixations','gyhduration', 'gyhm_duration','gyhsd_duration', 'gyhfixa','gyhfixb', 'gyhfixc', 'gyhfixd', 'gyhfixe',
             'gyhdura', 'gyhdurb', 'gyhdurc', 'gyhdurd', 'gyhdure','gyhmdisimg', 'gyhsddisimg', 'gyhmdisscan', 'gyhsddisscan',
             'gyhrevisits','gyhfirstfixationtime','gyhstd','gyhsasd','gyhstdt','gyhsasdt','gyh_degree_centrality_A',
           'gyh_degree_centrality_B','gyh_degree_centrality_C','gyh_degree_centrality_D','gyh_degree_centrality_E','gyh_betweenness_centrality_A'
           ,'gyh_betweenness_centrality_B','gyh_betweenness_centrality_C','gyh_betweenness_centrality_D','gyh_betweenness_centrality_E','gyh_eigenvector_centrality_A'
           ,'gyh_eigenvector_centrality_B','gyh_eigenvector_centrality_C','gyh_eigenvector_centrality_D','gyh_eigenvector_centrality_E','gyh_closeness_centrality_A'
           ,'gyh_closeness_centrality_B','gyh_closeness_centrality_C','gyh_closeness_centrality_D','gyh_closeness_centrality_E']
    for i in range(len(line1)):
        sheet3.write(0,i,line1[i])
    for i in range(len(features)):
        final=[]
        final.extend(features[i][:20])
        for j in range(len(firsttime)):
            if firsttime[j][0]==features[i][0]:
                final.append(firsttime[j][1])
                gyhfirsttime=firsttime[j][2]
                break
        for j in range(len(similarity)):
            if similarity[j][0]==features[i][0]:
                final.extend(similarity[j][1:5])
                gyhsimilarity=similarity[j][5:9]
                break
        for j in range(len(network)):
            if network[j][0]==features[i][0]:
                final.extend(network[j][1:21])
                gyhnetwork=network[j][21:]
                break
        final.extend(features[i][20:39])
        final.append(gyhfirsttime)
        final.extend(gyhsimilarity)
        final.extend(gyhnetwork)
        for j in range(len(final)):
            sheet3.write(i+1,j,final[j])
    f.save(f'{paramas.stanetfeatures}/media{file}.xls')


pa=paramas()
for i in range(1,15):
    joint(pa,i)



